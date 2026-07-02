#!/usr/bin/env python3
"""
CEMI Archival PDF Processor
============================
Converts scanned Soviet personnel records (АРАН Ф.1959 — ЦЭМИ АН СССР)
into multi-sheet Excel workbooks using Claude API for OCR and structuring.

USAGE:
  python cemi_processor.py [directory]

  - If directory is omitted, scans the current working directory.
  - The script lists all *.pdf files containing "1959" in the filename.
  - User selects ONE file to process.
  - User is prompted for ANTHROPIC_API_KEY (once per PDF).
  - Output: cemi_<delo>_<year>_full.xlsx in the same directory as the PDF.

DESIGN:
  - One PDF processed per invocation (re-run for next PDF).
  - Two-stage Claude pipeline:
    1. Per-page OCR (vision) → preserves Russian text, handwritten/struck
    2. Document-wide structuring → multi-sheet JSON schema
  - Excel build replicates manually-curated style:
    _Summary, _CrossReferences, List_*, Roster_*, etc.
    + 5-color highlighting (director / doctorates / KPSS / star figures / handwritten)
  - Per-page OCR cached in .cemi_cache/<pdf_stem>/ to allow resumption on failure.

DEPENDENCIES:
  pip install anthropic openpyxl PyMuPDF Pillow

AUTHOR: Generated assistant tooling for CEMI archive processing.
"""

import argparse
import base64
import getpass
import json
import os
import re
import sys
import textwrap
import time
from pathlib import Path
from typing import Any

# ---------- Dependencies ----------
_DEPS_OK = True
try:
    import anthropic
except ImportError:
    print("Missing dependency: pip install anthropic", file=sys.stderr)
    _DEPS_OK = False
try:
    import fitz  # PyMuPDF
except ImportError:
    print("Missing dependency: pip install PyMuPDF", file=sys.stderr)
    _DEPS_OK = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    _DEPS_OK = False
if not _DEPS_OK:
    sys.exit(1)


# ============================================================
# Constants & embedded knowledge base
# ============================================================

DEFAULT_MODEL = "claude-opus-4-5"  # change if a newer / cheaper model is preferred
DEFAULT_DPI = 150
MAX_TOKENS_OCR = 4096
MAX_TOKENS_STRUCTURE = 16384

# Color palette (matches manually-built workbooks)
PALETTE = {
    "header_fill":   "DDDDDD",
    "highlight":     "FFFFCC",  # yellow — star figures
    "director":      "C8E6C9",  # green — director / Nemchinov / Federenko
    "kpss":          "FFE0E0",  # pink — KPSS members
    "doctorate":     "E1F5FE",  # blue — doctors / candidates
    "handwritten":   "FFE6CC",  # orange — handwritten additions
    "wave1":         "FFE6CC",  # ИКТП wave (1963)
    "wave2":         "E1F5FE",  # ИЭ АН СССР wave
    "wave3":         "F0E6F6",  # СОПС wave (1963)
    "struck":        "EEEEEE",  # gray — struck-through entries
    "promotion":     "E1F5FE",  # light blue — special promotions (Shatalin)
}

# Highlight rule keys -> palette
COLOR_MAP = {
    "yellow":  PALETTE["highlight"],
    "green":   PALETTE["director"],
    "pink":    PALETTE["kpss"],
    "blue":    PALETTE["doctorate"],
    "orange":  PALETTE["handwritten"],
    "gray":    PALETTE["struck"],
    "purple":  PALETTE["wave3"],
}

# This goes into the Claude API system prompt for the structuring stage.
CEMI_DOMAIN_CONTEXT = """\
You are processing personnel records from the archive of the Central Economic-Mathematical Institute of the USSR Academy of Sciences (ЦЭМИ АН СССР), fond 1959, opis' 1.

INSTITUTIONAL HISTORY
- 1958-1960: a sector/group inside Институт экономики АН СССР, with the earliest documented staff tenure starting in January 1958.
- July 1960: Лаборатория по применению математических методов в экономических исследованиях и планировании АН СССР is constituted under Director academician В.С. НЕМЧИНОВ (1894-1964), located at ул. Волхонка 14, Moscow. Often abbreviated as "Лаборатория экономико-математических методов АН СССР".
- 1 October 1963: ЦЭМИ is established by Order of the Presidium of the AS USSR № 2-1577. Director: Н.П. ФЕДОРЕНКО (member-correspondent, then full academician 1964). Институт changes address to ул. Дм. Ульянова 3.
- November 1964: НЕМЧИНОВ dies; his laboratory is renamed in his honour.
- 1965: Зам.директора ОЛЕЙНИК-ОВОД defends his doctorate and is replaced by С.С. ШАТАЛИН (later architect of the "500 Days" plan).
- 1973: А.И. КАЦЕНЕЛИНБОЙГЕН emigrates to the USA.
- 1985: В.Л. МАКАРОВ becomes director.
- February 1986: mass transfer (~150 staff) to ИЭП НТП АН СССР during Gorbachev-era reorganisation.
- February 1987: nomenclature reform — bulk reclassification of academic ranks (главный научн.сотр., ведущий научн.сотр., старший научн.сотр., научн.сотр.).

KEY FIGURES (mark with stars; include in cross_references when present)
- ★★★ НЕМЧИНОВ В.С. (b.1894, d.5.XI.1964) — academician, founder, lab director through 1963; first ЦЭМИ zав.лабораторией.
- ★★★ ФЕДОРЕНКО Н.П. — ЦЭМИ founding director (1963-1985).
- ★★★ МАКАРОВ В.Л. — ЦЭМИ director from 10 July 1985.
- ★★★ ШАТАЛИН С.С. — joined 24 May 1965, promoted to зам.директора 15 October 1965; future "500 Days" reformer.
- ★★★ КАЦЕНЕЛИНБОЙГЕН А.И. — joined 1 IX 1963 from ИЭ АН СССР; emigrated USA 1973.
- ★★ ЛУРЬЕ А.Л. — defended doctoral 13.III.1963; linear-programming pioneer.
- ★★ МИНЦ Л.Е. (b.~1894, vuz 1914) — pre-revolution-trained statistician; doctoral defense 22.VI.1965.
- ★★ ВАЙНШТЕЙН А.Л. (b.1892) — pre-revolutionary statistician; ст.науч.сотр. title 25.XII.1964.
- ★★ ВИШНЕВ С.М. (b.1898) — long-tenured economist; retires 31.XII.1985.
- ★★★ СУВОРОВ Б.П. (b.1933) — earliest documented tenure (5.I.1958); appears in every packet through 1987.
- ★★ ОВСИЕНКО Ю.В. — joined 1.IX.1963 (ИЭ АН СССР wave); → 1987 главный научн.сотр., д.э.н.
- ★★ ФАЕРМАН Е.Ю. — joined 1.IX.1963; → 1987 ведущий научн.сотр., д.э.н.
- ★★ МЕДНИЦКИЙ В.Г. — joined 1.VII.1961; doctoral defense 23.IV.1985; → 1986/87 zав.лаб., д.э.н.
- ★★ ВОЛКОНСКИЙ В.А. — defended k.f.-m.n. at Steklov 1960; → 1987 главный научн.сотр., д.э.н.
- ★★ ГАВРИЛЕЦ Ю.Н. — joined 18.X.1960 (П/Я 2377); → 1987 zав.лаб., д.э.н., профессор.
- ★★ ДАДАЯН В.С. (Armenian) — joined 24.X.1960 with Nemchinov from СОПС; defended kandidat 12.III.1963; → 1986 главный научн.сотр., д.э.н.
- ★★ КОССОВ В.В. — joined 1.I.1959; defended kandidat 12.III.1963; full КПСС from January 1965; future Госплан deputy chair.
- ★ ГЛАЗЬЕВ С.Ю. — defended kandidat 13.X.1986; promoted 1.VI.1987.
- ★ ПОЛТЕРОВИЧ В.М. — first formal title 8.X.1987.
- ★ ИВАНТЕР, ЯРЕМЕНКО, ЛЕВАДА, ШОХИН, НАЙШУЛЬ, НЕЧАЕВ, САЛТЫКОВ, ЛОПУХИН, БЕЛАСОВ — depart with the February 1986 wave to ИЭП НТП.
- ★ АЙВАЗЯН С.А., ШЕВЯКОВ А.Ю. — promoted to deputy director in 1986.
- ★ АЛЕКСАШЕНКО С.В. — enters 1986 as stazher.

DOCUMENT FORM CODES BY ERA
- 1961: Form № 5-нк (annual statistical report) + Form № 9 (specialty / national / party breakdown).
- 1962: Form 10-НР (precursor of 10-НПР).
- 1963-1979: Form 10-НПР.
- 1980+: Form 5-нк (returned, but in modernised form).

LIST NUMBERING BY ERA
- 1962 (д.19): List_1=defenses, List_2=departures, List_3=inbound совмест., List_4=outbound совмест., КОНТРОЛЬНЫЙ СПИСОК=full roster.
- 1963 (д.30): List_1=defenses, List_2=departures, List_4=совмест., List_5=hires, List_5а=КПСС.
- 1965+ (д.76 onwards): List_1=hires, List_2=departures, List_3=КПСС, List_4=defenses (sometimes with List_4_supp = ВАК approvals), List_5=academic title awards, List_6/6а=совмест. in/out, List_7/7а=stazhers / stazher transitions.

DATE FORMAT CONVENTIONS
- Roman numeral months: I=01, II=02, III=03, IV=04, V=05, VI=06, VII=07, VIII=08, IX=09, X=10, XI=11, XII=12.
- Typewriter substitutions: УП = VII; УШ = VIII (Cyrillic capital U+П for Roman V+I).
- Slash format: "10/XII-64" → 1964-12-10.
- Dot format: "12.III-63 г." → 1963-03-12.
- Russian month names: января=01, февраля=02, …, декабря=12.

NAMING CONVENTIONS
- Family names are typically printed in CAPITALS in lists ("НЕМЧИНОВ В.С.").
- In cover pages of personnel cards (учётные карточки), names appear in genitive case ("Немчинова В.С.").
- "и.о." = исполняющий обязанности (acting); "к.э.н." = кандидат экономических наук; "д.э.н." = доктор экономических наук; "ст.науч.сотр." = старший научный сотрудник; "м.н.с." = младший научный сотрудник; "зав.лаб." = заведующий лабораторией.
- Address "ул.Д.Ульянова, 3" = ЦЭМИ from 1963 onwards.
- Address "Волхонка, 14" = pre-CEMI laboratory (1960-1963).

WORK PRODUCT REQUIREMENTS
1. Capture EVERY entry in EVERY list. Never summarise or skip.
2. Preserve exact Russian wording for names, organisations, dates.
3. Provide ISO-8601 dates alongside the verbatim form whenever computable.
4. Mark struck-through / handwritten entries explicitly.
5. Generate at least 8 substantive findings in summary_findings — each one a full paragraph that highlights a historiographical insight (cross-list patterns, era-specific quirks, biographical trajectories).
6. Cross-references must explain forward/backward trails to other CEMI packets where applicable.
7. Use star markings (★, ★★, ★★★) for figure importance in findings and cross-references.
"""


# ============================================================
# PDF rasterisation
# ============================================================

def rasterize_pdf(pdf_path: Path, image_dir: Path, dpi: int = DEFAULT_DPI) -> list[Path]:
    """Rasterise each page of the PDF to a JPEG; return list of paths in order."""
    image_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(str(pdf_path))
    paths: list[Path] = []
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    for i, page in enumerate(doc, start=1):
        out = image_dir / f"page_{i:03d}.jpg"
        if not out.exists():
            pix = page.get_pixmap(matrix=matrix)
            # Some packets are ZIP archives of JPEGs wrapped as PDF;
            # PyMuPDF still rasterises the rendered page correctly.
            pix.save(str(out), jpg_quality=85)
        paths.append(out)
    doc.close()
    return paths


def encode_image_b64(path: Path) -> str:
    return base64.standard_b64encode(path.read_bytes()).decode("ascii")


# ============================================================
# Anthropic API helpers
# ============================================================

def make_client(api_key: str) -> "anthropic.Anthropic":
    return anthropic.Anthropic(api_key=api_key)


def call_with_retries(callable_, *, retries: int = 3, base_delay: float = 4.0):
    """Naïve exponential-backoff wrapper for API rate limits / transient errors."""
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            return callable_()
        except Exception as exc:  # broad catch is acceptable for a CLI utility
            last_exc = exc
            wait = base_delay * (2 ** attempt)
            sys.stderr.write(f"  [retry {attempt+1}/{retries}] error: {exc}; sleeping {wait:.1f}s\n")
            time.sleep(wait)
    assert last_exc is not None
    raise last_exc


def ocr_page(client, image_path: Path, page_idx: int, total: int, model: str) -> str:
    """OCR one page using Claude vision and return verbatim Russian transcription."""
    image_b64 = encode_image_b64(image_path)
    prompt = textwrap.dedent(f"""\
        You are transcribing scanned page {page_idx} of {total} from a Soviet-era
        Russian archival document (АРАН Ф.1959 оп.1 — ЦЭМИ АН СССР personnel records).

        STRICT RULES:
        1. Transcribe ALL Russian text — typewritten and handwritten — preserving
           spelling, punctuation, layout structure, and abbreviations exactly.
        2. Use markdown tables when the page is tabular. Preserve column alignment.
        3. Preserve date formats with Roman-numeral months (e.g., "30.I-65 г.",
           "12/III-63 г.") — DO NOT convert to ISO here.
        4. For handwritten content, prefix the line with [HW] or wrap the segment
           as [HW: ...].
        5. For struck-through / crossed-out content, wrap as [STRUCK: ...].
        6. For uncertain readings (poor handwriting, smudged text), append [?]
           after the doubtful word.
        7. Note folio numbers in the upper-right corner as [folio: N] at the top.
        8. Note signatures at the bottom: [signed: <name>, <title>].
        9. Do NOT translate, summarise, or interpret. Verbatim transcription only.
        10. Identify list type if visible (e.g., "СПИСОК № 1", "Контрольный список",
            "Form № 5-нк") at the top of your output.

        Output format:
        ```
        [folio: <NN if visible>]
        [list type: <if identifiable>]

        <verbatim transcription with markdown tables, [HW], [STRUCK], [?] markers>

        [signed: ...]
        ```
        """)

    def _call():
        return client.messages.create(
            model=model,
            max_tokens=MAX_TOKENS_OCR,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": image_b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }],
        )

    response = call_with_retries(_call)
    text_parts = [block.text for block in response.content if block.type == "text"]
    return "\n".join(text_parts).strip()


def structure_document(client, ocr_pages: list[str], pdf_filename: str,
                       year_hint: str | None, model: str) -> dict[str, Any]:
    """Send all OCR'd pages and ask Claude to produce structured JSON."""
    joined = "\n\n".join(
        f"=== PAGE {i+1} ===\n{txt}"
        for i, txt in enumerate(ocr_pages)
    )

    user_prompt = textwrap.dedent(f"""\
        Filename: {pdf_filename}
        Year hint (from filename or context): {year_hint or "unknown"}

        Below is the verbatim OCR'd content of every page in the packet.
        Generate structured JSON that captures this document at the same level
        of historiographical depth that an expert archivist would produce.

        --- BEGIN OCR ---
        {joined}
        --- END OCR ---

        Required JSON schema:

        {{
          "metadata": {{
            "fond": "1959",
            "opis": "1",
            "delo": "<from filename or content>",
            "year": "<reporting year, YYYY>",
            "form_code": "<10-НПР | 10-НР | 5-нк | other>",
            "period_start": "<YYYY-MM-DD or null>",
            "period_end": "<YYYY-MM-DD or null>",
            "director": "<full name>",
            "director_title": "<академик / чл.-корр. / д.э.н. / etc.>",
            "deputies": ["<list>"],
            "secretary": "<academic secretary>",
            "hr_officer": "<HR officer>",
            "address": "<institute address>",
            "submission_date": "<YYYY-MM-DD or null>",
            "registration_number": "<e.g., 423/330-346>",
            "page_count": <int>,
            "folio_range": "<e.g., 25-33>"
          }},
          "lists": [
            {{
              "sheet_name": "<short, valid Excel sheet name e.g., List_1, Roster_1962, Form_5nk>",
              "title_ru": "<full Russian title from the document>",
              "title_en": "<English description>",
              "subtitle": "<context line, e.g., signing date, signatories>",
              "folio_range": "<e.g., 25-26>",
              "headers": ["<col 1>", "<col 2>", ...],
              "rows": [
                ["<r1c1>", "<r1c2>", ...],
                ...
              ],
              "iso_date_columns": [<0-based indices of columns with verbatim dates that should also be rendered as ISO>],
              "highlight_rules": [
                {{"match_substring_in_col": <0-based col index>, "substring": "<text>", "color": "yellow|green|pink|blue|orange|gray|purple"}}
              ]
            }}
          ],
          "summary_findings": [
            {{
              "title": "<English heading with leading ★/★★/★★★ for importance>",
              "body": "<English paragraph; may include Russian phrases and Cyrillic names>"
            }}
          ],
          "cross_references": [
            {{
              "entity": "<full name in Russian capitals>",
              "list_refs": "<which sheets/rows>",
              "born": "<year or em-dash>",
              "status": "<short Russian status line>",
              "trail": "<English forward/backward trail explanation>"
            }}
          ]
        }}

        QUALITY BAR:
        - Capture every list and every row. Never summarise or skip.
        - Provide at least 8 substantive findings (full paragraphs each).
        - Provide at least 6 cross-references for star figures present.
        - Include star prefixes (★, ★★, ★★★) where appropriate.
        - Highlight rules: at minimum mark director-level entries with "green",
          KPSS members with "pink", doctorate-holders with "blue",
          and major star figures with "yellow".
        - Use exact Russian wording from the OCR for names and dates.

        REPLY WITH VALID JSON ONLY. No markdown fences, no commentary.
        """)

    def _call():
        current_assistant_text = ""
        for _ in range(5):  # max 5 continuations
            messages = [{"role": "user", "content": user_prompt}]
            if current_assistant_text:
                messages.append({"role": "assistant", "content": current_assistant_text})
                
            response = client.messages.create(
                model=model,
                max_tokens=MAX_TOKENS_STRUCTURE,
                system=CEMI_DOMAIN_CONTEXT,
                messages=messages,
            )
            text = "".join(block.text for block in response.content if block.type == "text")
            current_assistant_text += text
            
            if response.stop_reason == "max_tokens":
                import sys
                sys.stdout.write("  [API] max_tokens reached, continuing generation...\n")
                sys.stdout.flush()
                continue
            else:
                break
        return current_assistant_text

    raw = call_with_retries(_call).strip()

    # Best-effort recovery if model wrapped JSON in fences or added prose
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    # Find the first '{' and last '}' as a fallback
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        raw = raw[start:end+1]

    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        # Save the raw text for inspection so the user can recover manually
        debug_path = Path.cwd() / "cemi_structure_raw.txt"
        debug_path.write_text(raw, encoding="utf-8")
        raise RuntimeError(
            f"Failed to parse structuring JSON: {exc}. "
            f"Raw response saved to {debug_path}"
        ) from exc


# ============================================================
# Excel rendering
# ============================================================

def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no [ ] : * ? / \\."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name)
    return cleaned[:31] or "Sheet"


def _roman_month_to_iso(verbatim: str | None) -> str | None:
    """Convert dd.ROM-yy / dd/ROM-yy / dd.ROM.yy / 'D <Russian month> YYYY' to ISO."""
    if verbatim is None:
        return None
    s = str(verbatim).strip()
    if not s or s in {"—", "-", '—"—', "нет"}:
        return None
    s = s.replace("г.", "").replace(" г", "").rstrip(".").strip()

    roman_map = {
        "I": "01", "II": "02", "III": "03", "IV": "04", "V": "05", "VI": "06",
        "VII": "07", "VIII": "08", "IX": "09", "X": "10", "XI": "11", "XII": "12",
        "УП": "07", "УШ": "08",
    }
    rus_months = {
        "января": "01", "февраля": "02", "марта": "03", "апреля": "04",
        "мая": "05", "июня": "06", "июля": "07", "августа": "08",
        "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12",
    }

    # Russian month name format: "5 января 1958"
    parts = s.split()
    if len(parts) >= 3:
        try:
            d = parts[0].zfill(2)
            m = rus_months.get(parts[1].lower())
            yr = parts[2]
            if m and yr.isdigit():
                if len(yr) == 2:
                    yr = "19" + yr
                return f"{yr}-{m}-{d}"
        except Exception:
            pass

    # Slash / dot / hyphen mixes
    candidate = s.replace("/", ".").replace("-", ".")
    chunks = [c for c in candidate.split(".") if c]
    if len(chunks) >= 3:
        try:
            d = chunks[0].strip().zfill(2)
            roman = chunks[1].strip().upper()
            yr = chunks[2].strip()
            if len(yr) == 2:
                yr = "19" + yr
            if roman in roman_map and yr.isdigit():
                return f"{yr}-{roman_map[roman]}-{d}"
        except Exception:
            pass
    return None


def _matches_highlight(value: Any, rule: dict) -> bool:
    if value is None:
        return False
    target = str(value).lower()
    sub = (rule.get("substring") or "").lower()
    return bool(sub) and sub in target


def build_excel(structured: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    title_font   = Font(name="Arial", size=12, bold=True)
    sub_font     = Font(name="Arial", size=10, italic=True)
    header_font  = Font(name="Arial", size=10, bold=True)
    body_font    = Font(name="Arial", size=10)
    header_fill  = PatternFill("solid", start_color=PALETTE["header_fill"])
    wrap         = Alignment(wrap_text=True, vertical="top")

    metadata = structured.get("metadata", {})
    lists    = structured.get("lists", [])
    findings = structured.get("summary_findings", [])
    xrefs    = structured.get("cross_references", [])

    # ---------------- _Summary ----------------
    ws = wb.create_sheet("_Summary")
    ws["A1"] = (
        f"ЦЭМИ АН СССР — Personnel Report "
        f"{metadata.get('year', '')} "
        f"(Ф.{metadata.get('fond','1959')} оп.{metadata.get('opis','1')} "
        f"д.{metadata.get('delo','?')})"
    )
    ws["A1"].font = title_font

    sub_lines = [
        f"Form {metadata.get('form_code','?')} | {metadata.get('page_count','?')} pages | "
        f"folios {metadata.get('folio_range','?')} | reporting period: "
        f"{metadata.get('period_start') or '?'} — {metadata.get('period_end') or '?'}",
        f"Director: {metadata.get('director_title','')} {metadata.get('director','')}".strip(),
    ]
    if metadata.get("deputies"):
        sub_lines.append("Deputies: " + ", ".join(metadata["deputies"]))
    if metadata.get("secretary"):
        sub_lines.append(f"Уч.секретарь: {metadata['secretary']}")
    if metadata.get("hr_officer"):
        sub_lines.append(f"Ст.инспектор по кадрам / Зав.ОК: {metadata['hr_officer']}")
    if metadata.get("address"):
        sub_lines.append(f"Address: {metadata['address']}")
    if metadata.get("submission_date") or metadata.get("registration_number"):
        sub_lines.append(
            f"Submission: {metadata.get('submission_date','')} "
            f"reg. № {metadata.get('registration_number','')}"
        )

    for i, line in enumerate(sub_lines, start=2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = sub_font

    # Sheet inventory
    inv_row = len(sub_lines) + 4
    ws.cell(row=inv_row, column=1, value="Sheet inventory").font = header_font
    headers = ["Sheet", "Title (RU)", "Description (EN)", "Rows", "Folio"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=inv_row + 2, column=j, value=h)
        c.font = header_font
        c.fill = header_fill

    inv_first = inv_row + 3
    total_rows = 0
    for i, lst in enumerate(lists):
        sheet = _safe_sheet_name(lst.get("sheet_name", f"List_{i+1}"))
        rows  = len(lst.get("rows", []))
        total_rows += rows
        for j, val in enumerate([
            sheet,
            lst.get("title_ru", ""),
            lst.get("title_en", ""),
            rows,
            lst.get("folio_range", ""),
        ], 1):
            cell = ws.cell(row=inv_first + i, column=j, value=val)
            cell.font = body_font
            cell.alignment = wrap

    total_row = inv_first + len(lists)
    ws.cell(row=total_row, column=1, value="TOTAL").font = header_font
    ws.cell(row=total_row, column=4, value=total_rows).font = header_font

    # Findings
    findings_start = total_row + 3
    ws.cell(row=findings_start, column=1, value="Key historiographical findings").font = header_font
    rr = findings_start + 2
    for f in findings:
        ws.cell(row=rr, column=1, value=f.get("title", "")).font = header_font
        body_cell = ws.cell(row=rr + 1, column=1, value=f.get("body", ""))
        body_cell.font = body_font
        body_cell.alignment = wrap
        rr += 2

    for j, w in enumerate([40, 60, 60, 12, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------------- _CrossReferences ----------------
    ws = wb.create_sheet("_CrossReferences")
    ws["A1"] = "Cross-references — multi-list entities and forward/backward trails"
    ws["A1"].font = title_font
    headers = ["Entity", "List refs", "Born", "Status (1962+)", "Forward / backward trail"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, x in enumerate(xrefs, start=4):
        for j, val in enumerate([
            x.get("entity", ""),
            x.get("list_refs", ""),
            x.get("born", ""),
            x.get("status", ""),
            x.get("trail", ""),
        ], 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = body_font
            cell.alignment = wrap
        ws.row_dimensions[i].height = 38

    for j, w in enumerate([34, 30, 10, 36, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------------- Data sheets ----------------
    for lst in lists:
        sheet_name = _safe_sheet_name(lst.get("sheet_name", "List"))
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = lst.get("title_ru", sheet_name)
        ws["A1"].font = title_font
        if lst.get("title_en"):
            ws["A2"] = lst["title_en"]
            ws["A2"].font = sub_font
        if lst.get("subtitle"):
            ws["A3"] = lst["subtitle"]
            ws["A3"].font = sub_font

        headers = list(lst.get("headers", []))
        rows    = list(lst.get("rows", []))
        iso_cols = lst.get("iso_date_columns") or []

        # Insert ISO columns after their source columns (rightward shift)
        # We instead append "(ISO)" columns at the right for clarity.
        derived_headers = list(headers)
        for col_idx in iso_cols:
            if 0 <= col_idx < len(headers):
                derived_headers.append(f"{headers[col_idx]} (ISO)")

        header_row = 5
        for j, h in enumerate(derived_headers, 1):
            c = ws.cell(row=header_row, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = wrap

        for i, row in enumerate(rows):
            row = list(row)  # ensure mutable
            # Pad / truncate to header length
            while len(row) < len(headers):
                row.append("")
            row = row[:len(headers)]
            # Append ISO conversions
            iso_extra = []
            for col_idx in iso_cols:
                if 0 <= col_idx < len(headers):
                    iso_extra.append(_roman_month_to_iso(row[col_idx]))
            full_row = row + iso_extra

            for j, val in enumerate(full_row, 1):
                cell = ws.cell(row=header_row + 1 + i, column=j, value=val)
                cell.font = body_font
                cell.alignment = wrap

            # Apply highlight rules
            for rule in lst.get("highlight_rules", []):
                col = rule.get("match_substring_in_col")
                color_key = rule.get("color", "yellow")
                fill_color = COLOR_MAP.get(color_key, PALETTE["highlight"])
                if isinstance(col, int) and 0 <= col < len(headers):
                    if _matches_highlight(row[col], rule):
                        for j in range(1, len(derived_headers) + 1):
                            ws.cell(row=header_row + 1 + i, column=j).fill = (
                                PatternFill("solid", start_color=fill_color)
                            )

        # Column widths: simple heuristic
        widths = []
        for h in derived_headers:
            if "ФИО" in h or "Фамилия" in h or "имя" in h.lower():
                widths.append(28)
            elif "Должность" in h or "должность" in h.lower():
                widths.append(28)
            elif "ISO" in h:
                widths.append(12)
            elif "№" in h:
                widths.append(6)
            elif "Дата" in h or "дата" in h.lower() or "Год" in h:
                widths.append(16)
            else:
                widths.append(34)
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[header_row].height = 36

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ============================================================
# Workflow
# ============================================================

def find_1959_pdfs(directory: Path) -> list[Path]:
    return sorted(p for p in directory.iterdir()
                  if p.is_file() and p.suffix.lower() == ".pdf" and "1959" in p.name)


def menu_select_pdf(pdfs: list[Path]) -> Path | None:
    print("\nPDFs found (filename contains '1959'):")
    for i, p in enumerate(pdfs, 1):
        size_kb = p.stat().st_size // 1024
        print(f"  [{i:2d}] {p.name}  ({size_kb} KB)")
    print(f"  [ 0] cancel")
    while True:
        raw = input("\nSelect a PDF by number: ").strip()
        if not raw:
            continue
        if raw == "0":
            return None
        if raw.isdigit() and 1 <= int(raw) <= len(pdfs):
            return pdfs[int(raw) - 1]
        print(f"  invalid; enter 1..{len(pdfs)} or 0 to cancel.")


def prompt_api_key() -> str:
    """Read the Anthropic API key (single prompt per run, i.e. per PDF)."""
    env_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if env_key:
        use_env = input("Use ANTHROPIC_API_KEY from environment? [Y/n] ").strip().lower()
        if use_env in {"", "y", "yes"}:
            return env_key
    while True:
        key = getpass.getpass("Enter ANTHROPIC_API_KEY (input hidden): ").strip()
        if key:
            return key
        print("  empty key; try again.")


def derive_year_hint(pdf_name: str) -> str | None:
    """Pull a 4-digit reporting year out of the filename if present.

    Filenames in the CEMI archive follow either:
      - "1959_1_<delo>.pdf"            (no year encoded; e.g. 1959_1_10.pdf)
      - "1959_1_<delo><year>.pdf"      (year suffix; e.g. 1959_1_761965.pdf → 1965)
      - "1959_1_<delo><year>_<n>.pdf"  (parted; e.g. 1959_1_7501979_1.pdf → 1979)

    We must skip the literal "1959_1_" prefix before searching for the year,
    otherwise the leading "1959" will always match.
    """
    stem = pdf_name.rsplit(".", 1)[0].replace(" ", "_").replace("(", "").replace(")", "")
    m = re.match(r"^1959_1_(.+)$", stem)
    body = m.group(1) if m else stem

    # Look for any 4-digit year 1959..2030 within the remaining body.
    candidates = re.findall(r"(19[5-9]\d|20[0-2]\d)", body)
    for cand in candidates:
        # Use the LAST plausible reporting year (most likely a suffix).
        year_int = int(cand)
        if 1959 <= year_int <= 2030:
            best = cand
    return candidates[-1] if candidates else None


def derive_output_path(pdf_path: Path, year_hint: str | None) -> Path:
    """cemi_<delo>_<year>_full.xlsx in the same directory as the PDF."""
    stem = pdf_path.stem.replace(" ", "_").replace("(", "").replace(")", "")
    # Try to pull delo number
    m = re.match(r"1959_1_(\d+?)(\d{4})?(?:_\d+)?$", stem)
    if m and m.group(1):
        delo = m.group(1)
    elif "_" in stem:
        # fallback: take chunk after second underscore
        parts = stem.split("_")
        delo = parts[2] if len(parts) >= 3 else stem
        delo = re.sub(r"\D+", "", delo) or stem
    else:
        delo = re.sub(r"\D+", "", stem) or stem

    year_part = f"_{year_hint}" if year_hint else ""
    out_name = f"cemi_1959_1_{delo}{year_part}_full.xlsx"
    return pdf_path.parent / out_name


def cache_dir_for(pdf_path: Path) -> Path:
    return pdf_path.parent / ".cemi_cache" / pdf_path.stem


def process_pdf(pdf_path: Path, model: str, dpi: int, force: bool) -> Path:
    print(f"\n=== Processing: {pdf_path.name} ===")
    cache = cache_dir_for(pdf_path)
    image_dir = cache / "images"
    ocr_dir   = cache / "ocr"
    ocr_dir.mkdir(parents=True, exist_ok=True)

    api_key = prompt_api_key()
    client  = make_client(api_key)

    # Step 1: rasterise
    print(f"  rasterising at {dpi} DPI …")
    images = rasterize_pdf(pdf_path, image_dir, dpi=dpi)
    print(f"  {len(images)} pages.")

    # Step 2: OCR each page (cached)
    ocr_pages: list[str] = []
    for i, img in enumerate(images, start=1):
        cache_file = ocr_dir / f"page_{i:03d}.txt"
        if cache_file.exists() and not force:
            print(f"  page {i:3d}: cached.")
            ocr_pages.append(cache_file.read_text(encoding="utf-8"))
            continue
        print(f"  page {i:3d}: OCR …", flush=True)
        text = ocr_page(client, img, i, len(images), model=model)
        cache_file.write_text(text, encoding="utf-8")
        ocr_pages.append(text)

    # Step 3: structure the document
    year_hint = derive_year_hint(pdf_path.name)
    print(f"  structuring {len(ocr_pages)} pages (year hint: {year_hint or 'none'}) …")
    structured = structure_document(client, ocr_pages, pdf_path.name, year_hint, model=model)
    (cache / "structured.json").write_text(
        json.dumps(structured, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # Step 4: build the workbook
    out_path = derive_output_path(pdf_path, year_hint)
    print(f"  writing {out_path} …")
    build_excel(structured, out_path)
    print(f"  done.")
    return out_path


# ============================================================
# CLI
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("directory", nargs="?", default=".",
                        help="directory to scan for *1959*.pdf (default: current dir)")
    parser.add_argument("--model", default=DEFAULT_MODEL,
                        help=f"Anthropic model id (default: {DEFAULT_MODEL})")
    parser.add_argument("--dpi", type=int, default=DEFAULT_DPI,
                        help=f"rasterisation DPI (default: {DEFAULT_DPI})")
    parser.add_argument("--force", action="store_true",
                        help="ignore cached OCR and re-OCR every page")
    args = parser.parse_args()

    directory = Path(args.directory).resolve()
    if not directory.is_dir():
        print(f"Not a directory: {directory}", file=sys.stderr)
        return 2

    pdfs = find_1959_pdfs(directory)
    if not pdfs and directory.name != "Raw Materials(Sample)":
        alt_dir = directory / "Raw Materials(Sample)"
        if alt_dir.is_dir():
            pdfs = find_1959_pdfs(alt_dir)
            if pdfs:
                directory = alt_dir

    if not pdfs:
        print(f"No PDFs containing '1959' in filename found in {directory}.")
        return 1

    selected = menu_select_pdf(pdfs)
    if selected is None:
        print("Cancelled.")
        return 0

    out_path = derive_output_path(selected, derive_year_hint(selected.name))
    if out_path.exists():
        ans = input(f"\nOutput already exists: {out_path}\n  overwrite? [y/N] ").strip().lower()
        if ans not in {"y", "yes"}:
            print("Cancelled.")
            return 0

    try:
        result = process_pdf(selected, model=args.model, dpi=args.dpi, force=args.force)
    except Exception as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        return 1

    print(f"\n✓ Saved: {result}")
    print("  Re-run the script (without arguments) to process another PDF.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
