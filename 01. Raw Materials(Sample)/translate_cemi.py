#!/usr/bin/env python3
"""
translate_cemi.py
==================

CEMI archival Excel files translator (Russian → English)

This script translates all Excel files matching "cemi_1959_1*.xlsx" in the
current directory using the Anthropic Claude API. It applies the same
translation conventions established for the CEMI 1959 archival database:

  - BGN/PCGN transliteration for personal/institutional names
  - Full English replacement of Cyrillic content
  - Standardized terminology for Soviet academic ranks, party status,
    institutional names, dates (Roman-numeral month notation)
  - Historical Russian terms preserved with English equivalents
    (soiskatel, aspirantura, kandidat, sovmestitel, etc.)
  - Excel structure preserved (formulas, multi-sheet layout, cell
    coordinates, formatting)

Usage
-----

Place this script in the folder containing the cemi_1959_1*.xlsx files
(or any parent directory) and run:

    python3 translate_cemi.py                # interactive: prompts for API key per file
    python3 translate_cemi.py --all          # process all files using one key prompt
    python3 translate_cemi.py FILE.xlsx      # process a single file
    python3 translate_cemi.py --dir PATH     # process files in a specific directory

Translated files are saved to the "EN/" subfolder with the suffix "_EN.xlsx".
The script automatically handles macOS NFC/NFD Unicode normalization.

Requirements
------------
  - Python 3.8+
  - openpyxl (pip install openpyxl)
  - anthropic (pip install anthropic)

Set ANTHROPIC_API_KEY environment variable to skip the per-file prompt.
"""

import argparse
import getpass
import glob
import json
import os
import re
import shutil
import sys
import time
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing 'openpyxl'. Install with:  pip install openpyxl")
    sys.exit(1)

try:
    from anthropic import Anthropic
except ImportError:
    print("Missing 'anthropic'. Install with:  pip install anthropic")
    sys.exit(1)


CYR = re.compile(r"[Ѐ-ӿԀ-ԯ]")
FILE_PATTERN = "cemi_1959_1*.xlsx"
OUTPUT_SUBDIR = "EN"
OUTPUT_SUFFIX = "_EN.xlsx"
DEFAULT_MODEL = "claude-opus-4-6"
BATCH_SIZE = 80
MAX_RETRIES = 3
RETRY_DELAY_BASE = 4


BGN = {
    'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'Yo','Ж':'Zh','З':'Z','И':'I',
    'Й':'Y','К':'K','Л':'L','М':'M','Н':'N','О':'O','П':'P','Р':'R','С':'S','Т':'T',
    'У':'U','Ф':'F','Х':'Kh','Ц':'Ts','Ч':'Ch','Ш':'Sh','Щ':'Shch','Ъ':"",'Ы':'Y','Ь':"",
    'Э':'E','Ю':'Yu','Я':'Ya',
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh','з':'z','и':'i',
    'й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
    'у':'u','ф':'f','х':'kh','ц':'ts','ч':'ch','ш':'sh','щ':'shch','ъ':"",'ы':'y','ь':"",
    'э':'e','ю':'yu','я':'ya',
}


def offline_transliterate(s):
    return ''.join(BGN.get(ch, ch) for ch in s)


TRANSLATION_SYSTEM_PROMPT = """\
You are translating cells of a Russian-language Soviet archival personnel database
(CEMI = Центральный экономико-математический институт АН СССР, 1958-1987) into
English for inclusion in a Scientific Data journal article.

You will receive a JSON array of {id, value} objects. Translate every Russian
cell to natural, accurate, fluent English. Return ONLY a JSON array of {id,
translation} objects -- no explanations, no markdown.

TRANSLATION CONVENTIONS (apply consistently):

1. Personal names: BGN/PCGN transliteration of Cyrillic to Latin.
   Examples: НЕМЧИНОВ -> NEMCHINOV, ВОЛКОНСКИЙ -> VOLKONSKY,
   ОЛЕЙНИК-ОВОД -> OLEYNIK-OVOD, КАЦЕНЕЛИНБОЙГЕН -> KATSENELINBOIGEN.
   Initials: СУВОРОВ Б.П. -> SUVOROV B.P.
   Full patronymic: Юрий Васильевич -> Yury Vasilyevich.

2. Soviet academic ranks (use these exact English forms):
   - академик -> Academician
   - чл.-корр.АН СССР -> Cor.Mem. AS USSR
   - доктор экономических наук / д.э.н. -> Doctor of Economic Sciences
   - кандидат экономических наук / к.э.н. -> Candidate of Economic Sciences
   - доктор/кандидат физ.-мат. наук / д./к.ф.-м.н. -> Doctor/Candidate of Phys.-Math. Sciences
   - доктор/кандидат технических наук / д./к.т.н. -> Doctor/Candidate of Technical Sciences
   - профессор -> professor; доцент -> docent
   - главный научный сотрудник -> chief researcher
   - ведущий научный сотрудник -> leading researcher
   - старший научный сотрудник / ст.науч.сотр. / ст.н.с. -> senior researcher
   - научный сотрудник -> researcher
   - младший научный сотрудник / мл.науч.сотр. / м.н.с. -> junior researcher
   - стажёр-исследователь -> research intern (stazher)
   - и.о. (исполняющий обязанности) -> Acting (e.g. И.о.зав.лаб. -> Acting Head of Laboratory)
   - зав.лабораторией / зав.лаб. -> Head of Laboratory
   - зав.отделом -> Head of Department; зав.сектором -> Head of Sector
   - зам.директора -> Deputy Director; учёный секретарь -> Academic Secretary
   - старший инженер / ст.инженер -> senior engineer
   - ведущий инженер / вед.инженер -> lead engineer
   - инженер -> engineer
   - старший лаборант -> senior laboratory assistant; лаборант -> laboratory assistant

3. Party / state status:
   - член КПСС / Чл.КПСС -> CPSU member
   - кандидат в члены КПСС -> CPSU candidate member
   - член ВЛКСМ -> VLKSM member (Komsomol)
   - беспартийный / б/п -> non-party (b/p)

4. Institutions:
   - АН СССР -> USSR AS (Academy of Sciences of the USSR)
   - ЦЭМИ -> TsEMI (or CEMI in scholarly contexts)
   - МГУ им.М.В.Ломоносова -> MGU named after M.V. Lomonosov
   - МГУ -> MGU; МАДИ -> MADI; МЭСИ -> MESI; МГИМО -> MGIMO; ВЗФЭИ -> VZFEI
   - Институт экономики АН СССР -> Institute of Economics, USSR AS
   - СОПС -> SOPS (Council for the Study of Productive Forces)
   - ИКТП Госплана СССР -> Institute of Complex Transport Problems, USSR Gosplan
   - Госплан -> Gosplan; Госбанк -> Gosbank; ВАК -> VAK; НИИ -> NII
   - Президиум АН СССР -> Presidium of USSR AS
   - п/я / а/я -> p.o.box; в/ч -> military unit

5. Dates: convert Roman-numeral month notation, keeping Roman numerals.
     30.I-65 г. -> 30.I.1965
     1.УШ-1965 -> 1.VIII.1965  (Cyrillic У=V, УП=VII, УШ=VIII, IУ=IV, ХП=XII)
     6.I-1965 г. -> 6.I.1965
   Keep already-ISO dates (1965-08-01) as-is.

6. Departure reasons:
   - Собственное желание -> Own request
   - Собственное желание (улучшение условий) -> Own request (improvement of conditions)
   - Переход на другую работу -> Transfer to other employment
   - По сокращению штатов -> Due to staff reduction
   - Исключён из списков в связи со смертью -> Removed from the lists due to death
   - Зачисление в аспирантуру -> Admission to aspirantura
   - Окончание срока аспирантуры -> Completion of the aspirantura term
   - Переход на пенсию по возрасту -> Transition to age-based retirement pension

7. Historical Russian terms -- transliterate AND give English in parentheses
   on first occurrence:
   - соискатель -> soiskatel (external candidate)
   - аспирантура -> aspirantura; аспирант -> aspirant (postgraduate student)
   - стажировка -> stazhirovka
   - совместительство / совместитель -> sovmestitelstvo / sovmestitel
   - кандидатский минимум -> kandidat minimum
   - партбилет -> party card; кандидатская карточка -> candidate card

8. Don't over-translate: soiskatel/aspirantura/sovmestitel/stazher are widely
   used in English-language Soviet historiography.

9. Sheet/list headers (СПИСОК № N -> LIST No. N or Control List No. N).

10. Preserve all numbers, ISO dates, English text, star markers (★, ★★, ★★★),
    file references (д.30, F.1959 op.1), folio references (фолио / л.).

11. Russian abbreviations:
    - г. (after a year) -> drop
    - р. -> r.; руб. -> rub.
    - д. -> d. (delo) or bldg. (in addresses)

12. MIXED content: translate only Russian portions; keep English as-is.

13. Partial / fragmentary cells:
      [STRUCK: XXXXX] -> [STRUCK: XXXXX]
      [HW] / [рукопись] -> [HW] / [handwritten]
      [нечётко] -> [unclear]

14. Output: clean scholarly English suitable for Scientific Data. Match the
    precise bureaucratic tone of the Soviet original.

OUTPUT FORMAT -- STRICT:
Return only a JSON array, e.g.:
  [{"id": 0, "translation": "Director of the Laboratory"}, ...]

Valid JSON, no trailing commas, no markdown code fences, no commentary.
Each id MUST match the input id. Provide a translation for EVERY input item,
even if the original is already English (echo it back unchanged).
"""


def to_nfd(p):
    return unicodedata.normalize("NFD", p)


def find_cemi_files(directory):
    directory = Path(directory)
    nfd_dir = to_nfd(str(directory))
    patterns = [
        os.path.join(nfd_dir, FILE_PATTERN),
        os.path.join(str(directory), FILE_PATTERN),
    ]
    found = {}
    for pat in patterns:
        for p in glob.glob(pat):
            found[Path(p).name] = Path(p)
    return sorted(p for name, p in found.items() if not name.endswith(OUTPUT_SUFFIX))


def output_path_for(src):
    base_dir = Path(to_nfd(str(src.parent)))
    out_dir = base_dir / OUTPUT_SUBDIR
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = src.stem + OUTPUT_SUFFIX
    return out_dir / out_name


def extract_translatable_cells(workbook):
    cells = []
    for sn in workbook.sheetnames:
        ws = workbook[sn]
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if isinstance(c.value, str):
                    if c.value.startswith("="):
                        continue
                    if not c.value.strip():
                        continue
                    if CYR.search(c.value):
                        cells.append((sn, c.coordinate, c.value))
    return cells


def apply_translations_to_workbook(workbook, translations):
    applied = 0
    for (sn, coord), new_value in translations.items():
        ws = workbook[sn]
        ws[coord] = new_value
        applied += 1
    return applied


def call_claude(client, model, batch, attempt=1):
    user_msg = (
        "Translate the following CEMI archival cells from Russian to English. "
        "Apply the conventions in the system prompt. Return only the JSON array.\n\n"
        + json.dumps(batch, ensure_ascii=False, indent=2)
    )
    try:
        resp = client.messages.create(
            model=model,
            max_tokens=8192,
            system=TRANSLATION_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )
    except Exception as e:
        if attempt < MAX_RETRIES:
            wait = RETRY_DELAY_BASE * (2 ** (attempt - 1))
            print(f"    API error ({type(e).__name__}): {e}. Retrying in {wait}s...")
            time.sleep(wait)
            return call_claude(client, model, batch, attempt + 1)
        raise

    text = "".join(blk.text for blk in resp.content if blk.type == "text").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        if attempt < MAX_RETRIES:
            print(f"    Invalid JSON, retrying...")
            time.sleep(RETRY_DELAY_BASE)
            return call_claude(client, model, batch, attempt + 1)
        raise RuntimeError(f"Could not parse Claude response as JSON: {e}\n{text[:500]}")
    if not isinstance(data, list):
        raise RuntimeError(f"Expected JSON array, got {type(data).__name__}")
    return data


def translate_file(src, api_key, model=DEFAULT_MODEL):
    print(f"\n{'='*70}")
    print(f"Processing: {src.name}")
    print(f"{'='*70}")

    out = output_path_for(src)
    print(f"  Output: {out}")
    if out.exists():
        ans = input(f"  Output already exists. Overwrite? [y/N]: ").strip().lower()
        if ans != "y":
            print("  Skipped.")
            return {"file": src.name, "status": "skipped"}

    print(f"  Copying source to output...")
    shutil.copy2(src, out)

    wb = openpyxl.load_workbook(out)
    print(f"  Sheets: {len(wb.sheetnames)}")

    print(f"  Extracting Cyrillic cells...")
    cells = extract_translatable_cells(wb)
    print(f"  Cells to translate: {len(cells)}")
    if not cells:
        print(f"  No Cyrillic content found. Done.")
        wb.save(out)
        return {"file": src.name, "status": "empty", "cells": 0}

    client = Anthropic(api_key=api_key)
    translations = {}

    n_batches = (len(cells) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"  Translating in {n_batches} batch(es) of up to {BATCH_SIZE} cells...")
    t0 = time.time()
    for i in range(0, len(cells), BATCH_SIZE):
        batch_cells = cells[i:i + BATCH_SIZE]
        batch_idx = i // BATCH_SIZE + 1
        batch = [{"id": j, "value": v} for j, (_, _, v) in enumerate(batch_cells)]
        print(f"    Batch {batch_idx}/{n_batches} ({len(batch)} cells)... ", end="", flush=True)
        try:
            results = call_claude(client, model, batch)
        except Exception as e:
            print(f"FAILED: {e}")
            print(f"    Falling back to offline transliteration for this batch.")
            results = [
                {"id": j, "translation": offline_transliterate(b["value"])}
                for j, b in enumerate(batch)
            ]
        result_by_id = {r["id"]: r["translation"] for r in results if "id" in r and "translation" in r}
        for j, (sn, coord, _) in enumerate(batch_cells):
            if j in result_by_id:
                translations[(sn, coord)] = result_by_id[j]
            else:
                translations[(sn, coord)] = offline_transliterate(batch_cells[j][2])
        print(f"done ({len(result_by_id)}/{len(batch)} mapped)")

    elapsed = time.time() - t0
    print(f"  All batches translated in {elapsed:.1f}s")

    print(f"  Applying translations to workbook...")
    applied = apply_translations_to_workbook(wb, translations)
    wb.save(out)
    print(f"  Applied: {applied} translations")

    wb2 = openpyxl.load_workbook(out)
    remaining = 0
    for sn in wb2.sheetnames:
        for row in wb2[sn].iter_rows():
            for c in row:
                if isinstance(c.value, str) and CYR.search(c.value) and not c.value.startswith("="):
                    remaining += 1
    print(f"  Remaining Cyrillic cells: {remaining}")

    return {
        "file": src.name,
        "status": "ok",
        "output": str(out),
        "cells": len(cells),
        "applied": applied,
        "remaining_cyrillic": remaining,
        "elapsed_s": round(elapsed, 1),
    }


def prompt_api_key(file_label="this file"):
    env_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if env_key:
        ans = input(
            f"\nFound ANTHROPIC_API_KEY in environment. Use it for {file_label}? [Y/n]: "
        ).strip().lower()
        if ans in ("", "y", "yes"):
            return env_key

    print(f"\n>>> Enter Anthropic API key for {file_label} <<<")
    print("    (input is hidden; paste your sk-ant-... key and press Enter)")
    key = getpass.getpass("API key: ").strip()
    if not key:
        print("Empty key; aborting.")
        sys.exit(1)
    return key


def main():
    ap = argparse.ArgumentParser(
        description="Translate cemi_1959_1*.xlsx files (Russian -> English) using Claude API.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("files", nargs="*",
                    help="Specific files to translate (default: scan current directory).")
    ap.add_argument("--dir", default=".",
                    help="Directory to scan for cemi_1959_1*.xlsx files (default: current).")
    ap.add_argument("--model", default=DEFAULT_MODEL,
                    help=f"Claude model (default: {DEFAULT_MODEL}).")
    ap.add_argument("--all", action="store_true",
                    help="Use one API key prompt for ALL files (instead of per-file).")
    ap.add_argument("--batch-size", type=int, default=80,
                    help="Cells per API request (default: 80).")
    args = ap.parse_args()

    global BATCH_SIZE
    BATCH_SIZE = args.batch_size

    if args.files:
        targets = [Path(f) for f in args.files]
        for t in targets:
            if not t.exists():
                print(f"File not found: {t}")
                sys.exit(1)
    else:
        targets = find_cemi_files(args.dir)
        if not targets:
            alt_dir = Path(args.dir) / "Raw Materials(Sample)"
            if alt_dir.is_dir():
                targets = find_cemi_files(alt_dir)

    if not targets:
        print(f"No files matching '{FILE_PATTERN}' found in {args.dir}.")
        sys.exit(1)

    print(f"\nFound {len(targets)} file(s) to translate:")
    for i, t in enumerate(targets, 1):
        print(f"  {i:2d}. {t.name}")

    needs_work = []
    already = []
    for t in targets:
        out = output_path_for(t)
        (already if out.exists() else needs_work).append(t)
    if already:
        print(f"\n{len(already)} already have outputs in EN/:")
        for t in already:
            print(f"  - {t.name}")

    print()

    api_key = None
    if args.all:
        api_key = prompt_api_key("ALL files in this run")

    results = []
    for i, t in enumerate(targets, 1):
        print(f"\n[{i}/{len(targets)}] {t.name}")
        if not args.all:
            api_key = prompt_api_key(t.name)
        try:
            res = translate_file(t, api_key, model=args.model)
            results.append(res)
        except KeyboardInterrupt:
            print("\nInterrupted.")
            break
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback; traceback.print_exc()
            results.append({"file": t.name, "status": "error", "error": str(e)})

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    ok = sum(1 for r in results if r.get("status") == "ok")
    skipped = sum(1 for r in results if r.get("status") == "skipped")
    errored = sum(1 for r in results if r.get("status") == "error")
    print(f"  Translated: {ok}")
    print(f"  Skipped:    {skipped}")
    print(f"  Errors:     {errored}")
    for r in results:
        sym = {"ok": "[OK]", "skipped": "[--]", "error": "[X]", "empty": "[..]"}.get(r.get("status"), "[?]")
        cells = r.get("cells", "?")
        rem = r.get("remaining_cyrillic", "")
        rem_str = f" (Cyrillic remaining: {rem})" if isinstance(rem, int) and rem > 0 else ""
        print(f"  {sym} {r['file']}: {r.get('status')} cells={cells}{rem_str}")


if __name__ == "__main__":
    main()
