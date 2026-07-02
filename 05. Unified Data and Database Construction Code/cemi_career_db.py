#!/usr/bin/env python3
"""
CEMI Career Database — builder
==============================================================================
A clean-room implementation of the strategy in *CEMI 통합 DB 구축 전략.docx*.
This file shares NO code with the earlier `cemi_unified_db.py` /
`cemi_v2_build.py` lineage; it is rewritten from scratch around three
demands stated by the project owner:

  (a) Every career field that appears in the source spreadsheets must be
      individually queryable in the database, not merged into free text.
  (b) When a researcher holds more than one position in a given year
      (e.g. *Acting Junior Researcher* and *Junior Researcher* in 1966),
      the career path must reflect the rank order *acting → substantive*,
      not the row order in the spreadsheet.
  (c) Dual employment (concurrent part-time post at an outside institution)
      must be classified as a separate career fact and never collapsed into
      the CEMI appointment timeline.

Run from anywhere, as long as the three input workbooks are findable:

    data.xlsx
    cemi_translation_db.xlsx
    research_field_subfield.xlsx

Resolution order for each file (and for the output DB):
  1. The path passed via --personnel / --translation / --research / --db
  2. The current working directory
  3. The directory of this script
"""

from __future__ import annotations

import argparse
import dataclasses as dc
import hashlib
import os
import re
import sqlite3
import sys
import textwrap
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterable, Iterator

import pandas as pd


# ══════════════════════════════════════════════════════════════════════════════
# §A·0  Institutional phase (CEMI vs predecessor lab)
# ══════════════════════════════════════════════════════════════════════════════
#
# CEMI was founded in 1963 by reorganisation of an earlier laboratory inside
# the USSR Academy of Sciences' Institute of Economics.  Records dated 1962
# (and any earlier observation years) come from that predecessor lab, the
# *Laboratory of Mathematical Methods Applied to Economic Research and
# Planning*.  We tag every year row in `calendar_year` with one of three
# phases so that the UI can show the institutional context unambiguously.

CEMI_FOUNDED_YEAR        = 1963
PREDECESSOR_FOUNDED_YEAR = 1958      # informational only; some life-event years are even older

PHASE_CEMI          = "CEMI"
PHASE_PREDECESSOR   = "Predecessor Lab"
PHASE_PRE_INSTITUTE = "Pre-Institute"

LABEL_CEMI          = "Central Economic Mathematical Institute"
LABEL_PREDECESSOR   = "Laboratory of Mathematical Methods Applied to Economic Research and Planning"
LABEL_PRE_INSTITUTE = "Pre-institute (life-event year, e.g. party-membership or BA-graduation date)"


def institute_phase_for(year):
    """Return (phase, label) for a given year.

    Years ≥ 1963 → CEMI proper.  1958–1962 → predecessor lab (the lab itself
    pre-existed 1958, but most institutional records start here).  Anything
    older is treated as `Pre-Institute`: such a year is never an institutional
    observation year — it can only show up because an individual life event
    (party membership, BA graduation, etc.) carried a pre-institutional date.
    """
    if year is None:
        return (PHASE_CEMI, LABEL_CEMI)
    if year >= CEMI_FOUNDED_YEAR:
        return (PHASE_CEMI, LABEL_CEMI)
    if year >= PREDECESSOR_FOUNDED_YEAR:
        return (PHASE_PREDECESSOR, LABEL_PREDECESSOR)
    return (PHASE_PRE_INSTITUTE, LABEL_PRE_INSTITUTE)


# ══════════════════════════════════════════════════════════════════════════════
# §A   Position rank ladder
# ══════════════════════════════════════════════════════════════════════════════
#
# Soviet research institutes used a layered hierarchy in which an "Acting"
# (исполняющий обязанности) prefix denotes a temporary placement *below* the
# substantive role.  The numbers below are dimensionless ordinal weights
# chosen so that an "Acting X" sits just under the substantive "X", and so
# that whole tiers can be compared with simple < / >.  Unrecognised positions
# fall through to a heuristic estimator (RANK_DEFAULT).

RANK_DEFAULT = 50.0

POSITION_RANK: dict[str, float] = {
    # apprentice tier
    "trainee":                          10,
    "research trainee":                 12,
    "graduate student":                 14,
    "aspirant":                         14,
    "stazhyor":                         12,
    "junior engineer":                  18,
    "engineer":                         22,
    "senior engineer":                  26,
    "lead engineer":                    30,
    # junior researcher tier  (acting just below substantive)
    "acting junior researcher":         33,
    "junior researcher":                36,
    # researcher tier
    "researcher":                       40,
    "senior scientific and technical researcher": 43,
    # senior researcher tier
    "acting senior researcher":         52,
    "senior researcher":                55,
    "senior researcher-consultant":     57,
    "senior researcher-consultant, professor": 58,
    "principal researcher":             59,
    # sector / lab head tier
    "acting head of sector":            62,
    "head of sector":                   65,
    "acting head of laboratory":        67,
    "head of laboratory":               70,
    # department / secretariat
    "deputy head of department":        73,
    "academic secretary":               75,
    "scientific secretary":             75,
    "head":                             77,
    "chief":                            77,
    # academic-title-as-position
    "associate professor":              60,
    "professor":                        80,
    # leadership
    "deputy director":                  92,
    "director":                         99,
    # admin/editorial
    "editor":                           38,
}


def position_rank(name: str | None) -> float:
    """Look up a numeric rank.  Falls through to a heuristic for unknown values."""
    if not name:
        return RANK_DEFAULT
    n = " ".join(name.strip().lower().split())
    if n in POSITION_RANK:
        return POSITION_RANK[n]
    # Heuristic — strip 'acting' and try again, then compose down by 3.
    if n.startswith("acting "):
        base = position_rank(n[len("acting "):])
        return base - 3.0
    # Fallbacks based on tokens
    if "director" in n:                return 95.0
    if "head"     in n:                return 70.0
    if "senior"   in n:                return 55.0
    if "lead"     in n:                return 30.0
    if "junior"   in n:                return 36.0
    if "engineer" in n:                return 22.0
    return RANK_DEFAULT


def is_acting(name: str | None) -> bool:
    return bool(name and name.strip().lower().startswith("acting "))


# ══════════════════════════════════════════════════════════════════════════════
# §B   Source-column constants  (kept here so a column rename only changes one place)
# ══════════════════════════════════════════════════════════════════════════════

SRC_PERSONNEL = {
    "year":           "year",
    "name":           "Name",
    "birth":          "Year of Birth",
    "ethnicity":      "Ethnicity",
    "party":          "Party Membership",
    "party_year":     "Year of Party Membership",
    "primary_inst":   "Primary Institution",
    "position":       "Position",
    "department":     "department",
    "start_date":     "Start Date of Work",
    "specialty":      "specialty, major field",
    "pre_inst":       "Previous Service Institution",
    "pre_type":       "Type of Previous Institutions",
    "pre_title":      "Previous Institution Title",
    "pt_inst":        "Name of the institution where employed part-time",
    "pt_type":        "Type of Part-time Institutions",
    "pt_position":    "position where employed part-time",
    "pt_duration":    "duration of the permitted part-time employment",
    "dismissal":      "date of dismissal",
    "transfer_inst":  "transferred institution",
    "transfer_type":  "Type of Transferred Institutions",
    "academic_title": "Academic Position",
    "degree":         "Degree",
    "phd_field":      "Field of Ph.D.",
    "diss_date":      "Dissertation defense date",
    "grad_school":    "Graduate School",
    "ba_school":      "Graduate School of BA",
}


# ══════════════════════════════════════════════════════════════════════════════
# §C   Normalisation primitives  (pure functions, unit-testable)
# ══════════════════════════════════════════════════════════════════════════════

DEGREE_CANON = {
    "doctors of sciences":      "National Doctor",
    "doctor of sciences":       "National Doctor",
    "doctors":                  "National Doctor",
    "candidates of sciences":   "Ph.D.",
    "candidate of sciences":    "Ph.D.",
    "candidates":               "Ph.D.",
    "national doctor":          "National Doctor",
    "ph.d.":                    "Ph.D.",
    "phd":                      "Ph.D.",
}

PARTY_CANON = {
    "member":                                ("CPSU", "Member"),
    "candidate":                             ("CPSU", "Candidate"),
    "member of cpsu":                        ("CPSU", "Member"),
    "member of the cpsu":                    ("CPSU", "Member"),
    "candidate of the cpsu":                 ("CPSU", "Candidate"),
    "candidate for cpsu membership":         ("CPSU", "Candidate"),
    "candidate for membership in the cpsu":  ("CPSU", "Candidate"),
    "member of the komsomol":                ("Komsomol", "Member"),
    "komsomol":                              ("Komsomol", "Member"),
}


def s(x: Any) -> str | None:
    """Trimmed string or None."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    out = str(x).strip()
    return out or None


def norm_text(x: Any) -> str | None:
    """Collapse whitespace inside a string."""
    v = s(x)
    return " ".join(v.split()) if v else None


def title_text(x: Any) -> str | None:
    v = norm_text(x)
    if not v:
        return None
    # Title-case but keep all-caps acronyms intact (e.g. CEMI, CPSU, USSR)
    out = []
    for tok in v.split():
        if re.fullmatch(r"[A-Z]{2,}", tok):
            out.append(tok)
        else:
            out.append(tok[0].upper() + tok[1:].lower() if tok else tok)
    return " ".join(out)


def split_name(raw: Any) -> tuple[str, str, str] | None:
    """Return (canonical, surname, initials) or None.

    Canonical is constructed so that **patronymic-distinct** persons
    (e.g. "Filippova Nina Alekseevna" vs "Filippova Natalya Arnoldovna")
    yield different canonical_names and therefore different person_ids:

      • Pure-initials raw  ("Surname I. O.")               → canonical "Surname I. O."
      • Fully-spelled raw  ("Surname Given Patronymic")    → canonical "Surname Given Patronymic"
      • Mixed             ("Surname Given. I." / partial) → canonical preserves the
                                                            spelled-out tokens verbatim

    A token is "spelled" when its alphabetic core is ≥ 3 characters
    (anything shorter is an initial — possibly with a trailing dot).
    The `initials` field is still derived consistently ("N. A.") so
    search-time pills and tag rows remain uniform across both styles.
    """
    v = s(raw)
    if not v:
        return None
    # Fast path — already strict initials form "Surname I.O."  Some 2-letter
    # initials (like "Ya.") are tolerated by the [A-Z][a-zA-Z]?\. pattern.
    if re.fullmatch(r"[A-Z][a-zA-Z'’\-]+(\s+[A-Z][a-zA-Z]?\.){1,3}", v):
        parts = v.split(None, 1)
        return v, parts[0], parts[1] if len(parts) > 1 else ""

    bare = v.replace(",", " ")
    parts = bare.split()
    if not parts:
        return None
    surname = parts[0][0].upper() + parts[0][1:].lower()
    rest_tokens = parts[1:]

    def _core(t: str) -> str:        # alphabetic body, no trailing punctuation
        c = t.rstrip(".").strip()
        for ch in "'’-":
            c = c.replace(ch, "")
        return c

    def _is_spelled(t: str) -> bool: # 3+ alpha chars in core → likely a given-name / patronymic
        c = _core(t)
        return len(c) >= 3 and c.isalpha()

    # Always derive the initials abbreviation for the `initials` field.
    initials_parts = []
    for t in rest_tokens:
        core = _core(t)
        if core:
            initials_parts.append(core[0].upper() + ".")
    initials = " ".join(initials_parts)

    any_spelled = any(_is_spelled(t) for t in rest_tokens)
    if any_spelled:
        # Preserve spelled tokens verbatim (Capitalise); shorten pure-
        # initial tokens to their letter so mixed forms read cleanly.
        canon_parts = []
        for t in rest_tokens:
            core = _core(t)
            if not core:
                continue
            if _is_spelled(t):
                canon_parts.append(core[0].upper() + core[1:].lower())
            else:
                canon_parts.append(core[0].upper() + ".")
        canon = f"{surname} {' '.join(canon_parts)}".strip()
    else:
        # Pure initials path.
        canon = f"{surname} {initials}".strip()
    return canon, surname, initials


def norm_degree(x: Any) -> str | None:
    v = s(x)
    if not v:
        return None
    return DEGREE_CANON.get(v.lower(), v)


def parse_party(x: Any) -> tuple[str, str] | None:
    v = s(x)
    if not v:
        return None
    return PARTY_CANON.get(v.lower())


def parse_iso(x: Any) -> str | None:
    """Normalise a date-ish value to ISO 'YYYY-MM-DD' (or 'YYYY-01-01' for bare years)."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if hasattr(x, "strftime"):
        return x.strftime("%Y-%m-%d")
    v = str(x).strip()
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", v)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    m = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", v)
    if m:
        d, mo, y = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    m = re.fullmatch(r"\d{4}", v)
    if m:
        return f"{v}-01-01"
    return v          # leave odd values as-is


def parse_int(x: Any) -> int | None:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return None


def parse_float(x: Any) -> float | None:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def hash_id(*parts: Any, prefix_chars: int = 12) -> int:
    """Deterministic ID — shared helper for everything that needs a stable PK."""
    payload = "│".join("" if p is None else str(p).strip().lower() for p in parts)
    h = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    return int(h[:prefix_chars], 16)


# ══════════════════════════════════════════════════════════════════════════════
# §D   Schema
# ══════════════════════════════════════════════════════════════════════════════

DDL = textwrap.dedent("""
    PRAGMA foreign_keys = ON;

    -- ── Identity ──────────────────────────────────────────────────────────
    CREATE TABLE IF NOT EXISTS person (
        person_id      INTEGER PRIMARY KEY,
        canonical_name TEXT NOT NULL UNIQUE,
        surname        TEXT NOT NULL,
        initials       TEXT,
        birth_year     INTEGER,
        ethnicity_id   INTEGER REFERENCES ethnicity(ethnicity_id)
    );
    CREATE TABLE IF NOT EXISTS person_name_variant (
        variant_id  INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id   INTEGER NOT NULL REFERENCES person(person_id),
        raw_name    TEXT NOT NULL,
        first_seen_year INTEGER,
        UNIQUE(person_id, raw_name)
    );

    -- ── Controlled vocabularies ───────────────────────────────────────────
    CREATE TABLE IF NOT EXISTS ethnicity (
        ethnicity_id INTEGER PRIMARY KEY,
        label_en     TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS calendar_year (
        year             INTEGER PRIMARY KEY,
        plan_period      TEXT,
        institute_phase  TEXT NOT NULL DEFAULT 'CEMI',
        institute_label  TEXT
    );
    CREATE TABLE IF NOT EXISTS position_rank (
        position_id  INTEGER PRIMARY KEY,
        label_en     TEXT NOT NULL UNIQUE,
        label_ru     TEXT,
        is_acting    INTEGER NOT NULL DEFAULT 0,
        rank_order   REAL    NOT NULL,
        tier         TEXT
    );
    CREATE TABLE IF NOT EXISTS academic_title (
        title_id    INTEGER PRIMARY KEY,
        label_en    TEXT NOT NULL UNIQUE,
        rank_order  REAL    NOT NULL
    );
    CREATE TABLE IF NOT EXISTS department (
        department_id INTEGER PRIMARY KEY,
        label         TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS institution (
        institution_id INTEGER PRIMARY KEY,
        label          TEXT NOT NULL UNIQUE,
        kind           TEXT,
        -- v7: enrichment from "institution glossary (Nolting 4-sector …)"
        --     workbook.  label_ru carries the canonical Russian short or
        --     long form; classification_evidence preserves the rationale
        --     used to assign `kind` (e.g. "RU: 'НИИ'" or "Manual: AON
        --     pri TsK KPSS — Academy of Social Sciences (VUZ-equivalent)").
        label_ru               TEXT,
        classification_evidence TEXT
    );
    CREATE TABLE IF NOT EXISTS specialty (
        specialty_id INTEGER PRIMARY KEY,
        label        TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS phd_field (
        field_id   INTEGER PRIMARY KEY,
        label      TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS school (
        school_id INTEGER PRIMARY KEY,
        label     TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS degree (
        degree_id INTEGER PRIMARY KEY,
        label     TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS party_org (
        party_id INTEGER PRIMARY KEY,
        label    TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS source_row (
        source_row_id  INTEGER PRIMARY KEY AUTOINCREMENT,
        sheet          TEXT NOT NULL,
        row_index      INTEGER NOT NULL,
        UNIQUE(sheet, row_index)
    );

    -- ── Per-row observation (1:1 with personnel sheet rows) ───────────────
    -- This is the “raw evidence” layer.  Everything else can be re-derived
    -- from these rows together with the lookups above.
    CREATE TABLE IF NOT EXISTS person_year_observation (
        observation_id   INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        year             INTEGER NOT NULL REFERENCES calendar_year(year),
        source_row_id    INTEGER NOT NULL UNIQUE REFERENCES source_row(source_row_id),
        primary_institution_id INTEGER REFERENCES institution(institution_id),
        department_id    INTEGER REFERENCES department(department_id),
        specialty_id     INTEGER REFERENCES specialty(specialty_id),
        degree_id        INTEGER REFERENCES degree(degree_id),
        academic_title_id INTEGER REFERENCES academic_title(title_id),
        grad_school_id   INTEGER REFERENCES school(school_id),
        ba_school_id     INTEGER REFERENCES school(school_id),
        notes            TEXT
    );

    -- ── Career events ─────────────────────────────────────────────────────
    -- Every "fact" in a researcher's career goes into one of these tables.
    -- `within_year_order` is the field that satisfies the requirement
    -- "Acting Junior Researcher must come before Junior Researcher in 1966".

    CREATE TABLE IF NOT EXISTS cemi_appointment (
        appointment_id     INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id          INTEGER NOT NULL REFERENCES person(person_id),
        year               INTEGER NOT NULL REFERENCES calendar_year(year),
        position_id        INTEGER NOT NULL REFERENCES position_rank(position_id),
        rank_order         REAL    NOT NULL,
        is_acting          INTEGER NOT NULL DEFAULT 0,
        within_year_order  INTEGER NOT NULL,
        observation_id     INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, year, position_id)
    );

    CREATE TABLE IF NOT EXISTS dual_position (
        dual_id          INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        year             INTEGER NOT NULL REFERENCES calendar_year(year),
        institution_id   INTEGER NOT NULL REFERENCES institution(institution_id),
        position_label   TEXT,
        duration_text    TEXT,
        institution_kind TEXT,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, year, institution_id, position_label)
    );

    CREATE TABLE IF NOT EXISTS pre_cemi_role (
        pre_id           INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        institution_id   INTEGER NOT NULL REFERENCES institution(institution_id),
        institution_kind TEXT,
        title            TEXT,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, institution_id, title)
    );

    -- ── Career-exit event ordering convention ────────────────────────
    -- When a person has BOTH a dismissal_event and a transfer_event in
    -- the same calendar year (e.g. Svetlov K. S., 1980), the causal order
    -- is fixed by the UI timeline renderer:
    --      1. dismissal_event   (person leaves CEMI)
    --      2. transfer_event    (person arrives at destination institution)
    -- The DB itself does not store an explicit intra-year order for these
    -- two event types; the UI's openPerson() applies the rule above so
    -- 555 such overlapping (person, year) pairs render in causal order.
    CREATE TABLE IF NOT EXISTS transfer_event (
        transfer_id      INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        year             INTEGER NOT NULL REFERENCES calendar_year(year),
        institution_id   INTEGER NOT NULL REFERENCES institution(institution_id),
        institution_kind TEXT,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, year, institution_id)
    );

    CREATE TABLE IF NOT EXISTS dismissal_event (
        dismissal_id     INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        year             INTEGER NOT NULL REFERENCES calendar_year(year),
        dismissal_date   DATE,
        position_id      INTEGER REFERENCES position_rank(position_id),
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, year, dismissal_date)
    );

    CREATE TABLE IF NOT EXISTS party_event (
        party_event_id   INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        party_id         INTEGER NOT NULL REFERENCES party_org(party_id),
        role             TEXT CHECK(role IN ('Member','Candidate') OR role IS NULL),
        join_year        INTEGER REFERENCES calendar_year(year),
        join_date        DATE,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, party_id, role, join_year)
    );

    CREATE TABLE IF NOT EXISTS phd_defense (
        defense_id       INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        field_id         INTEGER REFERENCES phd_field(field_id),
        defense_date     DATE,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, field_id, defense_date)
    );

    CREATE TABLE IF NOT EXISTS start_of_work (
        start_id         INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        start_date       DATE NOT NULL,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        -- v5: each start_of_work row is the start date of one position
        -- spell; span_id back-points to the cemi_position_span row that
        -- this start_date opens, so a single JOIN expresses the
        -- (person, date, position, span_order, n_years) context of every
        -- start event.  Back-filled in derive_career_artifacts() by
        -- matching (person_id, year(start_date)) against
        -- cemi_position_span.start_year.  NULL only if the source
        -- carries a start_date that does not land on the first year of
        -- any span (rare; preserved verbatim).
        span_id          INTEGER REFERENCES cemi_position_span(span_id),
        -- v5 semantic (redefined from earlier "earliest row per person"):
        -- 1 = this row begins the FIRST position spell (span_order = 1)
        -- and is therefore the actual CEMI / lab joining date.
        -- 0 = a later position-change start date (the source records
        -- promotions / role transitions as further start_of_work rows
        -- without a source-side disambiguator from the initial hire —
        -- see §3 codebook).  When the source did NOT record the
        -- original-hire row, EVERY start_of_work row of that person
        -- carries is_initial_join = 0; the earliest available row is
        -- a position-change date, not the actual hire.
        is_initial_join  INTEGER NOT NULL DEFAULT 0,
        UNIQUE(person_id, start_date)
    );

    CREATE TABLE IF NOT EXISTS academic_title_award (
        award_id         INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id        INTEGER NOT NULL REFERENCES person(person_id),
        title_id         INTEGER NOT NULL REFERENCES academic_title(title_id),
        first_year       INTEGER,
        observation_id   INTEGER REFERENCES person_year_observation(observation_id),
        UNIQUE(person_id, title_id)
    );

    -- ── Aggregate facts (year × dimension) ────────────────────────────────
    CREATE TABLE IF NOT EXISTS demo_personnel_totals (
        year                  INTEGER PRIMARY KEY REFERENCES calendar_year(year),
        total_scientists      INTEGER,
        total_all_staff       REAL,
        national_doctor       REAL,
        phd                   REAL,
        women                 REAL,
        women_national_doctor REAL,
        women_phd             REAL,
        source_material       TEXT     -- v5: provenance string from workbook
    );
    CREATE TABLE IF NOT EXISTS demo_academic_degrees (
        year             INTEGER PRIMARY KEY REFERENCES calendar_year(year),
        total_scientists REAL,
        national_doctor  REAL,
        phd              REAL,
        pct_with_degree  REAL,
        professors       REAL,
        docents           REAL,
        sns_title         REAL,
        academicians      REAL,
        source_material   TEXT     -- v5: provenance string from workbook
    );
    CREATE TABLE IF NOT EXISTS demo_position (
        year         INTEGER NOT NULL REFERENCES calendar_year(year),
        position_id  INTEGER NOT NULL REFERENCES position_rank(position_id),
        total            INTEGER,
        national_doctor  REAL,
        phd              REAL,
        cpsu_members     REAL,
        komsomol         REAL,
        source_material  TEXT,    -- v5: provenance string from workbook
        PRIMARY KEY (year, position_id)
    );
    CREATE TABLE IF NOT EXISTS nationality (
        nationality_id INTEGER PRIMARY KEY,
        label_ru       TEXT,
        label_en       TEXT NOT NULL UNIQUE
    );
    CREATE TABLE IF NOT EXISTS demo_nationality (
        year           INTEGER NOT NULL REFERENCES calendar_year(year),
        nationality_id INTEGER NOT NULL REFERENCES nationality(nationality_id),
        total           INTEGER,
        national_doctor REAL,
        phd             REAL,
        PRIMARY KEY (year, nationality_id)
    );
    CREATE TABLE IF NOT EXISTS age_bracket (
        bracket_id  INTEGER PRIMARY KEY,
        label_ru    TEXT,
        label_en    TEXT NOT NULL UNIQUE,
        lower_bound INTEGER,
        upper_bound INTEGER
    );
    CREATE TABLE IF NOT EXISTS demo_age (
        year       INTEGER NOT NULL REFERENCES calendar_year(year),
        bracket_id INTEGER NOT NULL REFERENCES age_bracket(bracket_id),
        total           INTEGER,
        national_doctor REAL,
        phd             REAL,
        -- v4: per-(year, bracket) academic position / title breakdown
        --  Source: cemi_translation_db.xlsx · "Age Distribution" sheet.
        --  Each cell is a count of persons in that age bracket who held the
        --  corresponding academic title or research-staff rank.
        academician_or_corr  REAL,
        professor            REAL,
        associate_professor  REAL,
        senior_researcher    REAL,
        junior_researcher    REAL,
        PRIMARY KEY (year, bracket_id)
    );
    CREATE TABLE IF NOT EXISTS research_field (
        rfield_id   INTEGER PRIMARY KEY,
        code        TEXT,
        label_ru    TEXT,
        label_en    TEXT NOT NULL,
        is_aggregate INTEGER NOT NULL DEFAULT 0,
        UNIQUE(code, label_en)
    );
    CREATE TABLE IF NOT EXISTS demo_field (
        year       INTEGER NOT NULL REFERENCES calendar_year(year),
        rfield_id  INTEGER NOT NULL REFERENCES research_field(rfield_id),
        total           INTEGER,
        national_doctor REAL,
        phd             REAL,
        PRIMARY KEY (year, rfield_id)
    );
    CREATE TABLE IF NOT EXISTS research_subfield (
        subfield_id INTEGER PRIMARY KEY,
        rfield_id   INTEGER REFERENCES research_field(rfield_id),
        label_en    TEXT NOT NULL,
        UNIQUE(rfield_id, label_en)
    );
    CREATE TABLE IF NOT EXISTS demo_subfield (
        year                  INTEGER NOT NULL REFERENCES calendar_year(year),
        period                TEXT    NOT NULL,            -- "YYYY.MM" snapshot label
        subfield_id           INTEGER NOT NULL REFERENCES research_subfield(subfield_id),
        total_personnel       REAL,
        national_doctor       REAL,
        phd                   REAL,
        extras_json           TEXT,
        schema_year_signature TEXT NOT NULL,
        PRIMARY KEY (year, period, subfield_id)
    );
    CREATE TABLE IF NOT EXISTS demo_party (
        year             INTEGER PRIMARY KEY REFERENCES calendar_year(year),
        total_scientists REAL, cpsu_members REAL, cpsu_pct REAL,
        komsomol REAL, komsomol_pct REAL,
        source_material  TEXT      -- v5: provenance string from workbook
    );
    CREATE TABLE IF NOT EXISTS demo_trainees (
        year                   INTEGER PRIMARY KEY REFERENCES calendar_year(year),
        total_trainees         INTEGER,
        from_other_institutes  INTEGER
    );
    CREATE TABLE IF NOT EXISTS glossary (
        glossary_id INTEGER PRIMARY KEY,
        russian     TEXT,
        english     TEXT,
        category    TEXT
    );
    CREATE TABLE IF NOT EXISTS sheets_index (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        year        INTEGER, institution TEXT,
        source_file TEXT, sheet_name TEXT,
        first_cell_ru TEXT, first_cell_en TEXT,
        rows INTEGER, cols INTEGER,
        role        TEXT,
        target_table TEXT
    );

    -- ── Derived: cemi_position_span ────────────────────────────────────────
    -- Each row is one continuous stretch during which a person held the
    -- same (position_id, is_acting).  Built by walking cemi_appointment in
    -- (year, within_year_order) order per person and folding consecutive
    -- rows that agree on (position_id, is_acting).  Populated by
    -- derive_career_artifacts() AFTER load_personnel; cleared & rebuilt
    -- on every --reset.
    CREATE TABLE IF NOT EXISTS cemi_position_span (
        span_id     INTEGER PRIMARY KEY AUTOINCREMENT,
        person_id   INTEGER NOT NULL REFERENCES person(person_id),
        position_id INTEGER NOT NULL REFERENCES position_rank(position_id),
        is_acting   INTEGER NOT NULL DEFAULT 0,
        start_year  INTEGER NOT NULL,
        end_year    INTEGER NOT NULL,
        n_years     INTEGER NOT NULL,
        span_order  INTEGER NOT NULL    -- 1-based order within the person's career
    );

    -- ── Derived: person_degree ─────────────────────────────────────────────
    -- One row per (person, degree).  first_year is the earliest year that
    -- person appears in person_year_observation with that degree, and
    -- (last_year, n_years) describe the span over which the degree is
    -- observed.  Populated by derive_career_artifacts() at build time;
    -- cleared & rebuilt on every --reset.
    CREATE TABLE IF NOT EXISTS person_degree (
        person_id    INTEGER NOT NULL REFERENCES person(person_id),
        degree_id    INTEGER NOT NULL REFERENCES degree(degree_id),
        first_year   INTEGER NOT NULL,
        last_year    INTEGER NOT NULL,
        n_years      INTEGER NOT NULL,
        -- v4: provenance of where / when the degree was earned.  All
        -- nullable because the source data is incomplete (39–40% of
        -- degree holders have a recorded graduate school; defense_date
        -- and PhD field exist only for some phd_defense rows).
        school_id    INTEGER REFERENCES school(school_id),
        defense_date DATE,
        field_id     INTEGER REFERENCES phd_field(field_id),
        PRIMARY KEY (person_id, degree_id)
    );

    -- ── Indexes for the career-search workload ────────────────────────────
    CREATE INDEX IF NOT EXISTS idx_obs_person   ON person_year_observation(person_id);
    CREATE INDEX IF NOT EXISTS idx_obs_year     ON person_year_observation(year);
    CREATE INDEX IF NOT EXISTS idx_appt_person  ON cemi_appointment(person_id);
    CREATE INDEX IF NOT EXISTS idx_appt_year    ON cemi_appointment(year);
    CREATE INDEX IF NOT EXISTS idx_appt_pos     ON cemi_appointment(position_id);
    CREATE INDEX IF NOT EXISTS idx_span_person  ON cemi_position_span(person_id);
    CREATE INDEX IF NOT EXISTS idx_span_pos     ON cemi_position_span(position_id);
    CREATE INDEX IF NOT EXISTS idx_sow_span     ON start_of_work(span_id);
    CREATE INDEX IF NOT EXISTS idx_pdeg_person ON person_degree(person_id);
    CREATE INDEX IF NOT EXISTS idx_dual_person  ON dual_position(person_id);
    CREATE INDEX IF NOT EXISTS idx_dual_year    ON dual_position(year);
    CREATE INDEX IF NOT EXISTS idx_dual_inst    ON dual_position(institution_id);
    CREATE INDEX IF NOT EXISTS idx_pre_person   ON pre_cemi_role(person_id);
    CREATE INDEX IF NOT EXISTS idx_pre_inst     ON pre_cemi_role(institution_id);
    CREATE INDEX IF NOT EXISTS idx_trans_person ON transfer_event(person_id);
    CREATE INDEX IF NOT EXISTS idx_trans_year   ON transfer_event(year);
    CREATE INDEX IF NOT EXISTS idx_dis_person   ON dismissal_event(person_id);
    CREATE INDEX IF NOT EXISTS idx_party_person ON party_event(person_id);
    CREATE INDEX IF NOT EXISTS idx_phd_person   ON phd_defense(person_id);
    CREATE INDEX IF NOT EXISTS idx_pnv_person   ON person_name_variant(person_id);
""")


# ══════════════════════════════════════════════════════════════════════════════
# §E   Builder runtime
# ══════════════════════════════════════════════════════════════════════════════

class Vocab:
    """In-memory lookup that mirrors a controlled-vocabulary table.

    The builder caches "label → ID" so we only emit one INSERT per distinct
    label, and the rest of the loaders can do a hash-map lookup instead of
    a SQL round-trip.
    """

    def __init__(self, conn: sqlite3.Connection, table: str, label_col: str,
                 id_col: str, extra_cols: Iterable[str] = ()):
        self.conn = conn
        self.table = table
        self.label_col = label_col
        self.id_col = id_col
        self.extra_cols = tuple(extra_cols)
        self._cache: dict[str, int] = {}
        # Pre-warm with anything already in the DB so the cache survives
        # multiple builder runs.
        cur = conn.execute(f"SELECT {id_col}, {label_col} FROM {table}")
        for i, lab in cur.fetchall():
            if lab is not None:
                self._cache[lab] = i

    def get(self, label: str | None, **extras: Any) -> int | None:
        if not label:
            return None
        if label in self._cache:
            return self._cache[label]
        new_id = hash_id(self.table, label)
        cols = [self.id_col, self.label_col, *self.extra_cols]
        vals = [new_id, label, *(extras.get(c) for c in self.extra_cols)]
        placeholders = ",".join("?" * len(cols))
        self.conn.execute(
            f"INSERT OR IGNORE INTO {self.table}({','.join(cols)}) VALUES({placeholders})",
            vals,
        )
        self._cache[label] = new_id
        return new_id


class Stage:
    """The build pipeline lives in a Stage.  Each phase is a method."""

    def __init__(self, db: Path):
        self.conn = sqlite3.connect(db)
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self.conn.executescript(DDL)
        self.stats: dict[str, int] = {}

        # Vocab handles
        self.eth   = Vocab(self.conn, "ethnicity",      "label_en",  "ethnicity_id")
        self.dept  = Vocab(self.conn, "department",     "label",     "department_id")
        self.inst  = Vocab(self.conn, "institution",    "label",     "institution_id", ["kind"])
        self.spec  = Vocab(self.conn, "specialty",      "label",     "specialty_id")
        self.fld   = Vocab(self.conn, "phd_field",      "label",     "field_id")
        self.scho  = Vocab(self.conn, "school",         "label",     "school_id")
        self.deg   = Vocab(self.conn, "degree",         "label",     "degree_id")
        self.party = Vocab(self.conn, "party_org",      "label",     "party_id")
        self.title = Vocab(self.conn, "academic_title", "label_en",  "title_id",      ["rank_order"])
        self.pos   = Vocab(self.conn, "position_rank",  "label_en",  "position_id",   ["is_acting", "rank_order", "tier"])
        self.year  = Vocab(self.conn, "calendar_year",  "year",      "year")  # treat as-if vocab
        self.nat   = Vocab(self.conn, "nationality",    "label_en",  "nationality_id", ["label_ru"])
        self.age   = Vocab(self.conn, "age_bracket",    "label_en",  "bracket_id",     ["label_ru", "lower_bound", "upper_bound"])
        self.rfield = Vocab(self.conn, "research_field", "label_en",  "rfield_id",      ["code", "label_ru", "is_aggregate"])
        self.rsub   = Vocab(self.conn, "research_subfield","label_en","subfield_id",    ["rfield_id"])

    # ── Vocab convenience wrappers ───────────────────────────────────────

    def ensure_year(self, y: int | None) -> int | None:
        if y is None:
            return None
        if y in self.year._cache:
            return y
        phase, label = institute_phase_for(int(y))
        self.conn.execute(
            "INSERT OR IGNORE INTO calendar_year(year, institute_phase, institute_label) "
            "VALUES(?,?,?)", (y, phase, label))
        self.year._cache[y] = y
        return y

    def ensure_position(self, raw: str | None) -> int | None:
        n = title_text(raw)
        if not n:
            return None
        if n in self.pos._cache:
            return self.pos._cache[n]
        rank = position_rank(n)
        acting = 1 if is_acting(n) else 0
        if rank < 30:
            tier = "apprentice"
        elif rank < 50:
            tier = "junior"
        elif rank < 65:
            tier = "senior"
        elif rank < 90:
            tier = "head"
        else:
            tier = "leadership"
        return self.pos.get(n, is_acting=acting, rank_order=rank, tier=tier)

    def ensure_title(self, raw: str | None) -> int | None:
        n = norm_text(raw)
        if not n:
            return None
        # Academic title rank is its own scale: Academician 100, Corresponding Member 80
        if "academician" in n.lower():
            r = 100.0
        elif "correspond" in n.lower():
            r = 80.0
        else:
            r = 50.0
        return self.title.get(n, rank_order=r)

    def ensure_inst(self, raw: str | None, kind: str | None = None) -> int | None:
        n = norm_text(raw)
        return self.inst.get(n, kind=norm_text(kind))

    # ── Source-row anchor ────────────────────────────────────────────────

    def register_source_row(self, sheet: str, row_index: int) -> int:
        cur = self.conn.execute(
            "INSERT INTO source_row(sheet, row_index) VALUES(?,?)", (sheet, row_index))
        return cur.lastrowid

    # ── Phase 1: personnel rows → people, observations, events ──────────

    def load_personnel(self, df: pd.DataFrame) -> None:
        print(f"\n   Loading personnel: {len(df):,} rows × {len(df.columns)} cols")

        # ── Pass A: collect person identities (so PersonNameVariant can
        #            include first_seen_year before we re-iterate)
        persons: dict[int, dict[str, Any]] = {}
        variants: dict[tuple[int, str], int] = {}
        for _, row in df.iterrows():
            split = split_name(row[SRC_PERSONNEL["name"]])
            if not split:
                continue
            canon, surname, initials = split
            pid = hash_id("person", canon)
            rec = persons.setdefault(pid, {
                "canon": canon, "surname": surname, "initials": initials,
                "birth": None, "ethnicity": None,
            })
            if rec["birth"] is None:
                v = parse_int(row[SRC_PERSONNEL["birth"]])
                if v: rec["birth"] = v
            if rec["ethnicity"] is None:
                v = norm_text(row[SRC_PERSONNEL["ethnicity"]])
                if v: rec["ethnicity"] = v
            raw = s(row[SRC_PERSONNEL["name"]])
            if raw:
                key = (pid, raw)
                yr = parse_int(row[SRC_PERSONNEL["year"]])
                if key not in variants or (yr and yr < variants[key]):
                    variants[key] = yr or 9999

        for pid, rec in persons.items():
            eth_id = self.eth.get(rec["ethnicity"]) if rec["ethnicity"] else None
            self.conn.execute(
                "INSERT OR IGNORE INTO person(person_id, canonical_name, surname, "
                "initials, birth_year, ethnicity_id) VALUES(?,?,?,?,?,?)",
                (pid, rec["canon"], rec["surname"], rec["initials"], rec["birth"], eth_id))
        for (pid, raw), yr in variants.items():
            self.conn.execute(
                "INSERT OR IGNORE INTO person_name_variant(person_id, raw_name, first_seen_year) "
                "VALUES(?,?,?)",
                (pid, raw, yr if yr != 9999 else None))
        self.stats["persons"] = len(persons)
        self.stats["name_variants"] = len(variants)

        # ── Pass B: per-row observations + events
        n_obs = n_appt = n_dual = n_pre = n_tr = n_dis = n_party = n_phd = n_start = n_award = 0
        seen_starts: set[tuple[int, str]] = set()
        seen_dis:    set[tuple[int, int, str | None]] = set()
        seen_party:  set[tuple[int, int, str | None, int | None]] = set()
        seen_phd:    set[tuple[int, int | None, str | None]] = set()
        seen_pre:    set[tuple[int, int, str | None]] = set()
        seen_dual:   set[tuple[int, int, int, str | None]] = set()
        seen_tr:     set[tuple[int, int, int]] = set()
        seen_appt:   set[tuple[int, int, int]] = set()
        seen_award:  set[tuple[int, int]] = set()

        for row_idx, row in df.iterrows():
            split = split_name(row[SRC_PERSONNEL["name"]])
            if not split:
                continue
            pid = hash_id("person", split[0])
            year = parse_int(row[SRC_PERSONNEL["year"]])
            self.ensure_year(year)
            src_id = self.register_source_row("CEMI Personnel Data", int(row_idx))

            # Observation
            obs_id = self._insert_observation(pid, year, src_id, row)
            n_obs += 1

            # Main CEMI appointment
            position_label = title_text(row[SRC_PERSONNEL["position"]])
            if position_label and year is not None:
                pos_id = self.ensure_position(position_label)
                if pos_id is not None and (pid, year, pos_id) not in seen_appt:
                    seen_appt.add((pid, year, pos_id))
                    self.conn.execute(
                        "INSERT INTO cemi_appointment(person_id, year, position_id, "
                        "rank_order, is_acting, within_year_order, observation_id) "
                        "VALUES(?,?,?,?,?,?,?)",
                        (pid, year, pos_id, position_rank(position_label),
                         1 if is_acting(position_label) else 0, 0, obs_id))
                    n_appt += 1

            # Dual position (concurrent post at OUTSIDE institution)
            pt_inst_label = norm_text(row[SRC_PERSONNEL["pt_inst"]])
            if pt_inst_label and year is not None:
                kind = norm_text(row[SRC_PERSONNEL["pt_type"]])
                inst_id = self.ensure_inst(pt_inst_label, kind=kind)
                role = norm_text(row[SRC_PERSONNEL["pt_position"]])
                key = (pid, year, inst_id, role)
                if key not in seen_dual:
                    seen_dual.add(key)
                    self.conn.execute(
                        "INSERT INTO dual_position(person_id, year, institution_id, "
                        "position_label, duration_text, institution_kind, observation_id) "
                        "VALUES(?,?,?,?,?,?,?)",
                        (pid, year, inst_id, role,
                         norm_text(row[SRC_PERSONNEL["pt_duration"]]),
                         kind, obs_id))
                    n_dual += 1

            # Pre-CEMI role.
            # The source spreadsheet uses Previous Service Institution = "CEMI"
            # to encode an in-CEMI promotion (e.g. Trainee → Junior Researcher),
            # not actual prior service.  Those 1,584 rows are already represented
            # by cemi_appointment / cemi_position_span — skip them here so that
            # pre_cemi_role stays semantically clean ("real" external prior
            # roles only).  We also tolerate the Russian acronym "ЦЭМИ" just in
            # case it shows up in future data.
            pre_inst_label = norm_text(row[SRC_PERSONNEL["pre_inst"]])
            pre_norm = (pre_inst_label or "").strip().upper()
            is_internal = pre_norm in {"CEMI", "ЦЭМИ"}  # ЦЭМИ
            if pre_inst_label and not is_internal:
                kind = norm_text(row[SRC_PERSONNEL["pre_type"]])
                inst_id = self.ensure_inst(pre_inst_label, kind=kind)
                title = norm_text(row[SRC_PERSONNEL["pre_title"]])
                key = (pid, inst_id, title)
                if key not in seen_pre:
                    seen_pre.add(key)
                    self.conn.execute(
                        "INSERT INTO pre_cemi_role(person_id, institution_id, "
                        "institution_kind, title, observation_id) VALUES(?,?,?,?,?)",
                        (pid, inst_id, kind, title, obs_id))
                    n_pre += 1
            elif pre_inst_label and is_internal:
                n_pre_skipped_internal = locals().get("n_pre_skipped_internal", 0) + 1
                # Hoist into enclosing scope so the build stats can report it.
                # (load_personnel uses bare locals for these counters.)
                # We mutate the function-local accumulator via a small dict pattern:
                self._pre_skip_n = getattr(self, "_pre_skip_n", 0) + 1

            # Transfer out
            tr_inst_label = norm_text(row[SRC_PERSONNEL["transfer_inst"]])
            if tr_inst_label and year is not None:
                kind = norm_text(row[SRC_PERSONNEL["transfer_type"]])
                inst_id = self.ensure_inst(tr_inst_label, kind=kind)
                key = (pid, year, inst_id)
                if key not in seen_tr:
                    seen_tr.add(key)
                    self.conn.execute(
                        "INSERT INTO transfer_event(person_id, year, institution_id, "
                        "institution_kind, observation_id) VALUES(?,?,?,?,?)",
                        (pid, year, inst_id, kind, obs_id))
                    n_tr += 1

            # Dismissal
            dis_date = parse_iso(row[SRC_PERSONNEL["dismissal"]])
            if dis_date and year is not None:
                pos_at = title_text(row[SRC_PERSONNEL["position"]])
                pos_id = self.ensure_position(pos_at) if pos_at else None
                key = (pid, year, dis_date)
                if key not in seen_dis:
                    seen_dis.add(key)
                    self.conn.execute(
                        "INSERT INTO dismissal_event(person_id, year, dismissal_date, "
                        "position_id, observation_id) VALUES(?,?,?,?,?)",
                        (pid, year, dis_date, pos_id, obs_id))
                    n_dis += 1

            # Party event
            parsed = parse_party(row[SRC_PERSONNEL["party"]])
            if parsed:
                party_label, role = parsed
                party_id = self.party.get(party_label)
                join_year = parse_int(row[SRC_PERSONNEL["party_year"]])
                if join_year is not None:
                    self.ensure_year(join_year)
                key = (pid, party_id, role, join_year)
                if key not in seen_party:
                    seen_party.add(key)
                    self.conn.execute(
                        "INSERT INTO party_event(person_id, party_id, role, "
                        "join_year, join_date, observation_id) VALUES(?,?,?,?,?,?)",
                        (pid, party_id, role, join_year,
                         f"{join_year:04d}-01-01" if join_year else None, obs_id))
                    n_party += 1

            # PhD defense
            field_label = norm_text(row[SRC_PERSONNEL["phd_field"]])
            diss_date = parse_iso(row[SRC_PERSONNEL["diss_date"]])
            if diss_date or field_label:
                field_id = self.fld.get(field_label) if field_label else None
                key = (pid, field_id, diss_date)
                if key not in seen_phd:
                    seen_phd.add(key)
                    self.conn.execute(
                        "INSERT INTO phd_defense(person_id, field_id, defense_date, "
                        "observation_id) VALUES(?,?,?,?)",
                        (pid, field_id, diss_date, obs_id))
                    n_phd += 1

            # Start of work
            start_date = parse_iso(row[SRC_PERSONNEL["start_date"]])
            if start_date and (pid, start_date) not in seen_starts:
                seen_starts.add((pid, start_date))
                self.conn.execute(
                    "INSERT INTO start_of_work(person_id, start_date, observation_id) "
                    "VALUES(?,?,?)", (pid, start_date, obs_id))
                n_start += 1

            # Academic title award (Academic Position column)
            title_label = norm_text(row[SRC_PERSONNEL["academic_title"]])
            if title_label:
                tid = self.ensure_title(title_label)
                key = (pid, tid)
                if key not in seen_award and tid is not None:
                    seen_award.add(key)
                    self.conn.execute(
                        "INSERT INTO academic_title_award(person_id, title_id, "
                        "first_year, observation_id) VALUES(?,?,?,?)",
                        (pid, tid, year, obs_id))
                    n_award += 1

        self.stats["observations"]      = n_obs
        self.stats["cemi_appointment"]  = n_appt
        self.stats["dual_position"]     = n_dual
        self.stats["pre_cemi_role"]     = n_pre
        skipped_internal = getattr(self, "_pre_skip_n", 0)
        if skipped_internal:
            self.stats["pre_cemi_skipped_internal"] = skipped_internal
            print(f"     [info] pre_cemi_role: skipped {skipped_internal} in-CEMI "
                  f"promotion rows (Previous Service Institution = 'CEMI').")
            self._pre_skip_n = 0  # reset for any future re-run on the same Stage
        self.stats["transfer_event"]    = n_tr
        self.stats["dismissal_event"]   = n_dis
        self.stats["party_event"]       = n_party
        self.stats["phd_defense"]       = n_phd
        self.stats["start_of_work"]     = n_start
        self.stats["academic_title_award"] = n_award

        # ── Pass C: derive within_year_order so a (year, person) with
        #            multiple appointments is ordered by rank ascending.
        self._compute_within_year_order()

        self.conn.commit()
        for k, v in self.stats.items():
            print(f"     {k:24s} {v:>8,}")

    def _insert_observation(self, pid: int, year: int | None, src_id: int,
                            row: pd.Series) -> int:
        primary_inst = self.ensure_inst(row[SRC_PERSONNEL["primary_inst"]])
        dept_id      = self.dept.get(norm_text(row[SRC_PERSONNEL["department"]]))
        spec_id      = self.spec.get(norm_text(row[SRC_PERSONNEL["specialty"]]))
        deg_id       = self.deg.get(norm_degree(row[SRC_PERSONNEL["degree"]]))
        title_id     = self.ensure_title(norm_text(row[SRC_PERSONNEL["academic_title"]]))
        grad_id      = self.scho.get(norm_text(row[SRC_PERSONNEL["grad_school"]]))
        ba_id        = self.scho.get(norm_text(row[SRC_PERSONNEL["ba_school"]]))
        cur = self.conn.execute(
            "INSERT INTO person_year_observation(person_id, year, source_row_id, "
            "primary_institution_id, department_id, specialty_id, degree_id, "
            "academic_title_id, grad_school_id, ba_school_id) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)",
            (pid, year, src_id, primary_inst, dept_id, spec_id, deg_id, title_id,
             grad_id, ba_id))
        return cur.lastrowid

    def _compute_within_year_order(self) -> None:
        """Number a person's appointments inside a given year by rank ascending.

        That is the rule the project owner formulated: in 1966 if a person
        is recorded both as *Acting Junior Researcher* (rank 33) and
        *Junior Researcher* (rank 36), the path is 33 → 36, hence
        within_year_order 1 → 2.
        """
        cur = self.conn.execute("""
            SELECT appointment_id, person_id, year, rank_order
              FROM cemi_appointment
              ORDER BY person_id, year, rank_order, appointment_id
        """)
        rows = cur.fetchall()
        last_pid, last_year = None, None
        order = 0
        updates: list[tuple[int, int]] = []
        for appt_id, pid, year, rank in rows:
            if (pid, year) != (last_pid, last_year):
                last_pid, last_year = pid, year
                order = 1
            else:
                order += 1
            updates.append((order, appt_id))
        self.conn.executemany(
            "UPDATE cemi_appointment SET within_year_order = ? WHERE appointment_id = ?",
            updates)

    # ── Phase 1.5: derive career artifacts ───────────────────────────────
    #
    # Populates two derived structures from the freshly-loaded
    # cemi_appointment and start_of_work tables:
    #   • cemi_position_span        — consecutive same-(position, is_acting)
    #                                  appointment rows folded into spans.
    #   • start_of_work.is_initial_join — set on the earliest start row per
    #                                  person (ties broken by lowest start_id).

    def derive_career_artifacts(self) -> None:
        print("\n   Deriving career artifacts (position spans + initial-join flag)")
        cur = self.conn.cursor()

        # ── (a) cemi_position_span ─────────────────────────────────────
        cur.execute("DELETE FROM cemi_position_span")
        appts = cur.execute("""
            SELECT person_id, year, within_year_order, position_id,
                   rank_order, is_acting
              FROM cemi_appointment
             ORDER BY person_id, year, within_year_order
        """).fetchall()

        spans = []
        cur_span = None
        last_pid = None
        span_order = 0
        for pid, yr, wyo, pos_id, ro, acting in appts:
            if pid != last_pid:
                if cur_span is not None:
                    spans.append(cur_span)
                cur_span = None
                last_pid = pid
                span_order = 0
            if cur_span is not None and cur_span["position_id"] == pos_id \
                    and cur_span["is_acting"] == acting:
                cur_span["end_year"] = yr
            else:
                if cur_span is not None:
                    spans.append(cur_span)
                span_order += 1
                cur_span = {
                    "person_id":  pid,
                    "position_id": pos_id,
                    "is_acting":  acting,
                    "start_year": yr,
                    "end_year":   yr,
                    "span_order": span_order,
                }
        if cur_span is not None:
            spans.append(cur_span)

        rows = [(s["person_id"], s["position_id"], s["is_acting"],
                 s["start_year"], s["end_year"],
                 s["end_year"] - s["start_year"] + 1,
                 s["span_order"]) for s in spans]
        cur.executemany("""
            INSERT INTO cemi_position_span
              (person_id, position_id, is_acting, start_year, end_year,
               n_years, span_order)
              VALUES(?,?,?,?,?,?,?)""", rows)
        self.stats["cemi_position_span"] = len(rows)
        print(f"     cemi_position_span        {len(rows):>8,}  "
              f"(folded from {len(appts):,} appointments)")

        # ── (b) start_of_work.span_id + redefined is_initial_join ─────
        # Each start_of_work row is the start date of one cemi_position_span.
        # span_id is back-filled by a three-tier match against
        # cemi_position_span:
        #   tier-1  exact: start_date YEAR  ==  cemi_position_span.start_year
        #                  (pick lowest span_order on ties).
        #   tier-2  pre-CEMI: start_date YEAR < the person's earliest span
        #                  start_year (e.g. a 1958 / 1959 / 1960 predecessor-
        #                  lab join while the first cemi_appointment is
        #                  in 1961).  Such rows link to span_order = 1
        #                  because they semantically open the same first
        #                  spell, just before CEMI's own records begin.
        #   tier-3  mid-span: start_date YEAR falls strictly inside a
        #                  span (start_year < y ≤ end_year) without
        #                  matching its start_year — these are within-
        #                  year position transitions that the year-grain
        #                  span fold absorbs into the surrounding spell;
        #                  the row links to that containing span but
        #                  carries is_initial_join = 0.
        # Any row that fails all three tiers retains span_id = NULL and
        # is preserved verbatim.
        #
        # is_initial_join is redefined STRICTLY as "this row begins the
        # first position spell (span_order = 1)".  Persons whose
        # original-hire row was not recorded by the source carry
        # is_initial_join = 0 on every start_of_work row — the earliest
        # available date is a position-change date, not the actual hire.
        cur.execute("UPDATE start_of_work SET span_id = NULL, is_initial_join = 0")

        # Build the lookups we need:
        #   • spans_by_year[(pid, sy)]   = [(span_order, span_id), ...] sorted by order
        #   • spans_by_person[pid]       = [(start_year, end_year, span_order, span_id), ...] sorted
        # The per-year list is ordered by span_order so that, when a person
        # has multiple spans whose start_year is the same (e.g. Altaev 1965
        # Senior Engineer #1 then Acting Junior Researcher #2), the order
        # carries the *temporal* sequence within the year — within_year_order
        # of cemi_appointment seeded that span_order.
        spans_by_year   = {}
        spans_by_person = {}
        for sp_id, pid, sy, ey, span_order in cur.execute("""
            SELECT span_id, person_id, start_year, end_year, span_order
              FROM cemi_position_span
        """):
            spans_by_year.setdefault((pid, sy), []).append((span_order, sp_id))
            spans_by_person.setdefault(pid, []).append((sy, ey, span_order, sp_id))
        for k in spans_by_year:    spans_by_year[k].sort()
        for k in spans_by_person:  spans_by_person[k].sort()

        # Group start_of_work by (person, year), ordered by start_date.
        # Multiple rows in the same year are paired with same-year spans
        # i-th start ↔ i-th span (by date / by span_order respectively).
        starts_by_year = {}
        for sid, pid, sdate in cur.execute("""
            SELECT start_id, person_id, start_date
              FROM start_of_work
             WHERE start_date IS NOT NULL
        """):
            try:
                yr = int(str(sdate)[:4])
            except (TypeError, ValueError):
                continue
            starts_by_year.setdefault((pid, yr), []).append((str(sdate), sid))
        for k in starts_by_year:   starts_by_year[k].sort()

        updates_span = []
        updates_init = []
        n_exact = n_pre = n_mid = n_null = n_initial = 0

        # Also account for date-less rows: we have to count them as NULL
        # later because they never enter starts_by_year.
        n_null_dateless = cur.execute(
            "SELECT COUNT(*) FROM start_of_work WHERE start_date IS NULL"
        ).fetchone()[0]

        # Initial-join candidates are collected per person; only the
        # chronologically earliest row that maps to span_order = 1 is
        # actually flagged.  This avoids double-flagging when both the
        # pre-CEMI tier (e.g. 1964-11-01) and the exact tier (1965-01-01)
        # hit span_order = 1 for the same person.
        init_cand = {}      # pid -> (earliest_start_date, start_id)
        for (pid, yr), starts_list in starts_by_year.items():
            same_year_spans = spans_by_year.get((pid, yr), [])
            plist           = spans_by_person.get(pid, [])
            for i, (sdate, sid) in enumerate(starts_list):
                if i < len(same_year_spans):
                    # tier-1 exact: pair the i-th start with the i-th span.
                    sord, spid = same_year_spans[i]
                    updates_span.append((spid, sid))
                    n_exact += 1
                    if sord == 1:
                        prev = init_cand.get(pid)
                        if prev is None or sdate < prev[0]:
                            init_cand[pid] = (sdate, sid)
                    continue

                # No more same-year span to pair with — try mid-span first,
                # then pre-CEMI / predecessor lab.
                hit = None
                if plist:
                    for sy, ey, sord, spid in plist:
                        if sy <= yr <= ey:
                            hit = (spid, sord, "mid")   # don't break — last wins
                    if hit is None:
                        earliest_sy = plist[0][0]
                        if yr < earliest_sy:
                            sy0, ey0, sord0, spid0 = plist[0]
                            hit = (spid0, sord0, "pre")
                if hit is None:
                    n_null += 1
                    continue
                spid, sord, tier = hit
                updates_span.append((spid, sid))
                if tier == "mid":
                    n_mid += 1
                else:  # pre
                    n_pre += 1
                    if sord == 1:
                        prev = init_cand.get(pid)
                        if prev is None or sdate < prev[0]:
                            init_cand[pid] = (sdate, sid)

        # Materialise the deduped initial-join updates.
        updates_init = [(sid,) for (_, sid) in init_cand.values()]
        n_initial    = len(updates_init)

        # Roll the date-less rows into the unmatched bucket for the log.
        n_null += n_null_dateless

        if updates_span:
            cur.executemany(
                "UPDATE start_of_work SET span_id = ? WHERE start_id = ?",
                updates_span)
        if updates_init:
            cur.executemany(
                "UPDATE start_of_work SET is_initial_join = 1 "
                "WHERE start_id = ?",
                updates_init)

        n_total_linked = n_exact + n_pre + n_mid
        self.stats["start_of_work (span_id exact)"] = n_exact
        if n_pre:  self.stats["start_of_work (span_id pre-CEMI)"]   = n_pre
        if n_mid:  self.stats["start_of_work (span_id mid-span)"]   = n_mid
        if n_null: self.stats["start_of_work (span_id NULL)"]       = n_null
        self.stats["start_of_work (initial joins)"]                  = n_initial
        print(f"     start_of_work.span_id    {n_total_linked:>9,}  "
              f"(exact {n_exact} + pre-CEMI {n_pre} + mid-span {n_mid}; "
              f"unmatched {n_null})")
        print(f"     start_of_work initial flag{n_initial:>8,}  "
              f"(rows opening span_order = 1; persons whose original "
              f"hire row is not recorded carry 0 on every row)")

        # ── (c) person_degree ──────────────────────────────────────────
        # One row per (person, degree).  Built by grouping on
        # person_year_observation.  Persons observed with no degree_id
        # do not appear here; persons holding both Ph.D. and National
        # Doctor get two rows.  Three nullable columns are then back-
        # filled where the source data carries the information:
        #   • school_id    — earliest non-NULL grad_school_id for that
        #                    (person, degree).
        #   • defense_date — for Ph.D., earliest non-NULL defense_date
        #                    from phd_defense.
        #   • field_id     — for Ph.D., the field associated with that
        #                    earliest defense row.
        cur.execute("DELETE FROM person_degree")
        cur.execute("""
            INSERT INTO person_degree(person_id, degree_id,
                                      first_year, last_year, n_years)
            SELECT person_id, degree_id,
                   MIN(year), MAX(year), COUNT(DISTINCT year)
              FROM person_year_observation
             WHERE year IS NOT NULL AND degree_id IS NOT NULL
             GROUP BY person_id, degree_id
        """)
        n_deg = cur.execute("SELECT COUNT(*) FROM person_degree").fetchone()[0]

        # Back-fill school_id: earliest non-NULL grad_school for that
        # (person, degree).  Walk the observations once and remember
        # only the first hit per key.
        school_map = {}     # (person_id, degree_id) -> school_id
        for pid, did, sid in cur.execute("""
            SELECT person_id, degree_id, grad_school_id
              FROM person_year_observation
             WHERE grad_school_id IS NOT NULL AND degree_id IS NOT NULL
               AND year IS NOT NULL
             ORDER BY person_id, degree_id, year, observation_id
        """):
            key = (pid, did)
            if key not in school_map:
                school_map[key] = sid
        if school_map:
            cur.executemany(
                "UPDATE person_degree SET school_id = ? "
                "WHERE person_id = ? AND degree_id = ?",
                [(sid, pid, did) for (pid, did), sid in school_map.items()])

        # Back-fill defense_date + field_id per (person, degree).
        #
        # The source spreadsheet has a single "Dissertation defense date"
        # column that captures BOTH Ph.D. and National Doctor defenses —
        # whichever the person earned.  We route each phd_defense row to
        # the correct person_degree row by matching the defense YEAR
        # against person_degree.first_year for that person.
        #
        # Rule:
        #   • If defense_year matches a (person, degree) first_year,
        #     attach defense_date + field_id to THAT degree row.
        #     Aivazyan S. A. 1976-12-10 → National Doctor (first_year=1976).
        #   • If defense_year is BEFORE all of the person's first_year
        #     entries, the defense pre-dates the CEMI record window.
        #     Attach to Ph.D. (the typical earliest degree).
        #   • Date-less phd_defense rows (field-only) attach to Ph.D. as
        #     fallback so the field annotation is not lost.
        phd_row = cur.execute(
            "SELECT degree_id FROM degree WHERE label = 'Ph.D.'").fetchone()
        n_def = n_field = 0
        if phd_row:
            phd_id = phd_row[0]

            # Build (person_id, first_year) → degree_id lookup
            deg_by_year = {}
            for pid, did, fy in cur.execute(
                "SELECT person_id, degree_id, first_year FROM person_degree"):
                deg_by_year[(pid, fy)] = did

            phd_rows = list(cur.execute(
                "SELECT person_id, defense_date, field_id FROM phd_defense"))

            # For each (person, target_degree) we want to keep:
            #   • the EARLIEST non-NULL defense_date
            #   • a non-NULL field_id if one is available
            picks = {}  # (pid, did) -> (defense_date, field_id)
            def _bump(pid, did, ddate, fid):
                key = (pid, did)
                prev = picks.get(key)
                if prev is None:
                    picks[key] = (ddate, fid)
                    return
                p_date, p_fid = prev
                new_date = p_date
                if ddate is not None and (p_date is None or ddate < p_date):
                    new_date = ddate
                new_fid = p_fid if p_fid is not None else fid
                picks[key] = (new_date, new_fid)

            for pid, ddate, fid in phd_rows:
                # Route the row.
                target_did = phd_id   # default = Ph.D.
                if ddate is not None:
                    try:
                        def_year = int(str(ddate)[:4])
                    except (TypeError, ValueError):
                        def_year = None
                    if def_year is not None:
                        # Exact year match — attach to that degree.
                        hit = deg_by_year.get((pid, def_year))
                        if hit is not None:
                            target_did = hit
                        # Otherwise keep default = Ph.D.
                _bump(pid, target_did, ddate, fid)

            for (pid, did), (ddate, fid) in picks.items():
                cur.execute(
                    "UPDATE person_degree SET defense_date = ?, field_id = ? "
                    "WHERE person_id = ? AND degree_id = ?",
                    (ddate, fid, pid, did))
                if ddate is not None: n_def += 1
                if fid   is not None: n_field += 1

        n_school = len(school_map)
        self.stats["person_degree"]              = n_deg
        self.stats["person_degree (with school)"] = n_school
        if n_def:   self.stats["person_degree (with defense_date)"] = n_def
        if n_field: self.stats["person_degree (with phd field)"]    = n_field
        print(f"     person_degree            {n_deg:>8,}  "
              f"(one row per person×degree)")
        print(f"     person_degree.school     {n_school:>8,}  "
              f"(rows back-filled with grad_school_id)")
        if n_def:
            print(f"     person_degree.defense    {n_def:>8,}  "
                  f"(PhD rows with defense_date)")
        if n_field:
            print(f"     person_degree.field      {n_field:>8,}  "
                  f"(PhD rows with field_id)")
        self.conn.commit()

    # ── Phase 2: aggregates ──────────────────────────────────────────────

    @staticmethod
    def _pick(row, *names):
        """Return the first non-NaN value among the given column names.
        Tolerates workbook column renames between revisions (e.g.
        "Doctors of Sciences" → "National Doctors").  Returns None
        if no name is found and present in the row.
        """
        for nm in names:
            if nm in row.index:
                v = row[nm]
                # pandas treats float NaN as missing; pd.isna also handles None
                try:
                    import pandas as _pd
                    if not _pd.isna(v):
                        return v
                except Exception:
                    if v is not None:
                        return v
        return None

    def load_aggregates(self, xl: dict[str, pd.DataFrame]) -> None:
        print("\n   Loading aggregates")

        # Personnel totals
        df = xl["Personnel Totals"]
        rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year"))
            if y is None: continue
            self.ensure_year(y)
            rows.append((y, parse_int(r.get("Total Scientists")),
                         parse_float(r.get("Total All Staff")),
                         parse_float(self._pick(r, "National Doctors", "Doctors of Sciences")),
                         parse_float(self._pick(r, "Ph.D.s", "Candidates of Sciences")),
                         parse_float(r.get("Women")),
                         parse_float(self._pick(r, "Women National Doctors", "Women Doctors")),
                         parse_float(self._pick(r, "Women Ph.D.s",           "Women Candidates")),
                         norm_text(self._pick(r, "Source Material", "Source Sheet"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_personnel_totals VALUES(?,?,?,?,?,?,?,?,?)", rows)
        self.stats["demo_personnel_totals"] = len(rows)

        # Academic degrees
        df = xl["Academic Degrees"]; rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year"))
            if y is None: continue
            self.ensure_year(y)
            tot = parse_float(r.get("Total Scientists"))
            nd  = parse_float(self._pick(r, "National Doctors", "Doctors of Sciences"))
            phd = parse_float(self._pick(r, "Ph.D.s",           "Candidates of Sciences"))
            pct = parse_float(r.get("% with Degree"))
            if pct is None and tot and nd is not None and phd is not None:
                pct = round(100 * (nd + phd) / tot, 1)
            rows.append((y, tot, nd, phd, pct,
                         parse_float(r.get("Professors")),
                         parse_float(r.get("Docents")),
                         parse_float(r.get("SNS Title")),
                         parse_float(r.get("Academicians")),
                         norm_text(r.get("Source Material"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_academic_degrees VALUES(?,?,?,?,?,?,?,?,?,?)", rows)
        self.stats["demo_academic_degrees"] = len(rows)

        # Positions (vocab + fact)
        df = xl["Positions"]; rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year")); en = title_text(r.get("Position (English)"))
            if y is None or not en: continue
            self.ensure_year(y)
            pid = self.ensure_position(en)
            # Persist the Russian label on the position dimension (only if not yet set
            # — the personnel side fills English first; this aggregate side fills RU).
            ru = norm_text(r.get("Position (Russian)"))
            if ru:
                self.conn.execute(
                    "UPDATE position_rank "
                    "SET label_ru = COALESCE(label_ru, ?) "
                    "WHERE position_id = ?",
                    (ru, pid))
            rows.append((y, pid, parse_int(r.get("Total")),
                         parse_float(self._pick(r, "National Doctors", "Doctors")),
                         parse_float(self._pick(r, "Ph.D.s",           "Candidates")),
                         parse_float(r.get("CPSU Members")),
                         parse_float(r.get("Komsomol")),
                         norm_text(r.get("Source Material"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_position VALUES(?,?,?,?,?,?,?,?)", rows)
        self.stats["demo_position"] = len(rows)

        # Nationality
        df = xl["Nationality"]; rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year")); en = norm_text(r.get("Nationality (English)"))
            if y is None or not en: continue
            self.ensure_year(y)
            ru = norm_text(r.get("Nationality (Russian)"))
            nid = self.nat.get(en, label_ru=ru)
            rows.append((y, nid, parse_int(r.get("Total")),
                         parse_float(self._pick(r, "National Doctors", "Doctors")),
                         parse_float(self._pick(r, "Ph.D.s",           "Candidates"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_nationality VALUES(?,?,?,?,?)", rows)
        self.stats["demo_nationality"] = len(rows)

        # Age distribution
        df = xl["Age Distribution"]; rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year")); en = norm_text(r.get("Age Bracket (English)"))
            if y is None or not en: continue
            self.ensure_year(y)
            ru = norm_text(r.get("Age Bracket (Russian)"))
            lo = up = None
            m = re.match(r"^(\d+)\s*[–-]\s*(\d+)", en)
            if m: lo, up = int(m.group(1)), int(m.group(2))
            m2 = re.match(r"^up to\s*(\d+)", en, re.I)
            if m2: up = int(m2.group(1))
            m3 = re.match(r"^(\d+)\s*and (older|over|above)", en, re.I)
            if m3: lo = int(m3.group(1))
            bid = self.age.get(en, label_ru=ru, lower_bound=lo, upper_bound=up)
            rows.append((y, bid, parse_int(r.get("Total")),
                         parse_float(self._pick(r, "National Doctors", "Doctors")),
                         parse_float(self._pick(r, "Ph.D.s",           "Candidates")),
                         parse_float(r.get("Academician or Corresponding Member")),
                         parse_float(r.get("Professor")),
                         parse_float(r.get("Associate Professor")),
                         parse_float(r.get("Senior Researcher")),
                         parse_float(r.get("Junior Researcher"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_age VALUES(?,?,?,?,?,?,?,?,?,?)", rows)
        self.stats["demo_age"] = len(rows)

        # Research fields (top-level) — the source for this data has moved
        # from cemi_translation_db.xlsx · "Research Fields" sheet to the
        # research_field_subfield.xlsx workbook.  We no longer attempt to
        # load this sheet (it may have been deleted from the translation
        # workbook).  See `derive_field_aggregates()` below — it runs after
        # `load_subfields()` and populates `demo_field` by summing the
        # snapshot subfields per (year, Research Field) using the latest
        # snapshot for each year.
        if "Research Fields" in xl:
            print("     [skip] 'Research Fields' sheet found in translation "
                  "workbook but ignored — superseded by "
                  "research_field_subfield.xlsx.")

        # Party membership
        df = xl["Party Membership"]; rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year"))
            if y is None: continue
            self.ensure_year(y)
            rows.append((y, parse_float(r.get("Total Scientists")),
                         parse_float(r.get("CPSU Members")), parse_float(r.get("CPSU %")),
                         parse_float(r.get("Komsomol")), parse_float(r.get("Komsomol %")),
                         norm_text(r.get("Source Material"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_party VALUES(?,?,?,?,?,?,?)", rows)
        self.stats["demo_party"] = len(rows)

        # Trainees
        df = xl["Research Trainees"]; rows = []
        for _, r in df.iterrows():
            y = parse_int(r.get("Year"))
            if y is None: continue
            self.ensure_year(y)
            rows.append((y, parse_int(r.get("Total Trainees")),
                         parse_int(r.get("From Other Institutes"))))
        self.conn.executemany(
            "INSERT OR REPLACE INTO demo_trainees VALUES(?,?,?)", rows)
        self.stats["demo_trainees"] = len(rows)

        # Glossary  ────────────────────────────────────────────────────
        # Defensive loader: locate the sheet (alias-tolerant), normalise
        # degree-name English so the academic-degrees facet stays consistent,
        # and synthesise a row id (#) when the workbook omits one.  Emits an
        # explicit warning if the sheet is missing or empty so that an empty
        # Glossary tab in the UI can be traced back to the source workbook.
        #
        # Column detection is intentionally promiscuous: we try a list of
        # explicit aliases first, then fall back to case-insensitive substring
        # matching, then — as a last resort — sniff cell content to spot the
        # Cyrillic and Latin columns.  This means a Russian-glossary workbook
        # that uses headers like "Original term" / "English term" / "RU" / "EN"
        # still loads correctly instead of silently producing blank cells.
        gloss_aliases = ["Translation Glossary", "translation glossary",
                         "Glossary", "glossary", "글로사리", "용어집"]
        gloss_key = next((k for k in gloss_aliases if k in xl), None)
        if gloss_key is None:
            # Last-ditch: any sheet whose name contains "glossary"
            gloss_key = next((k for k in xl
                              if "glossary" in str(k).lower()), None)
        if gloss_key is None:
            print("     [warn] No 'Translation Glossary' sheet found in "
                  f"translation workbook (sheets: {list(xl)[:6]}…). "
                  "Glossary tab will be empty.")
            self.stats["glossary"] = 0
        else:
            df = xl[gloss_key].copy()
            # Column aliases: tolerate header drift across versions of the
            # workbook (e.g. "Russian", "Russian (original)") so that an
            # accidental rename does not silently empty the glossary.
            col_aliases = {
                "id":       ["#", "id", "ID", "Id", "No", "no", "row", "번호"],
                "russian":  ["Russian (original)", "Russian", "russian",
                             "RU", "ru", "Original", "Original term",
                             "Term (RU)", "원문", "러시아어"],
                "english":  ["English (translation)", "English", "english",
                             "EN", "en", "Translation", "English term",
                             "Term (EN)", "번역", "영어"],
                "category": ["Category", "category", "Cat", "분류", "Kind",
                             "Type"],
            }
            resolved = {}
            for key, cands in col_aliases.items():
                resolved[key] = next((c for c in cands if c in df.columns),
                                     None)
            # Case-insensitive substring fallback for whichever fields are
            # still unresolved.  This catches "russian (original)" (lowercase
            # snapshot from older xlsx versions) or "Cat." with a trailing dot.
            def _fuzzy(predicates):
                for c in df.columns:
                    name = str(c).lower()
                    if any(p in name for p in predicates):
                        return c
                return None
            if resolved["russian"]  is None: resolved["russian"]  = _fuzzy(["russian", "ru", "원문", "러시"])
            if resolved["english"]  is None: resolved["english"]  = _fuzzy(["english", "en", "번역", "영어", "translation"])
            if resolved["category"] is None: resolved["category"] = _fuzzy(["category", "cat", "분류", "kind", "type"])
            if resolved["id"]       is None: resolved["id"]       = _fuzzy(["id", "no", "#", "번호"])
            # Content-based detection: if russian/english are still missing
            # but the sheet clearly has Cyrillic and Latin text columns, find
            # them by sampling each object column.  Beats failing silently.
            if resolved["russian"] is None or resolved["english"] is None:
                cyrillic_re = re.compile(r"[Ѐ-ӿ]")
                latin_re    = re.compile(r"[A-Za-z]")
                cyr_col = lat_col = None
                cyr_score = lat_score = 0
                for c in df.columns:
                    sample = df[c].dropna().astype(str).head(40).tolist()
                    if not sample:
                        continue
                    cyr = sum(bool(cyrillic_re.search(v)) for v in sample)
                    lat = sum(bool(latin_re.search(v)) for v in sample)
                    if cyr > cyr_score:
                        cyr_score, cyr_col = cyr, c
                    if lat > lat_score and (cyr == 0 or lat > cyr * 2):
                        lat_score, lat_col = lat, c
                if resolved["russian"] is None and cyr_col is not None:
                    resolved["russian"] = cyr_col
                if resolved["english"] is None and lat_col is not None and lat_col != resolved.get("russian"):
                    resolved["english"] = lat_col
            print(f"     [info] glossary columns resolved → "
                  f"id={resolved['id']!r}, ru={resolved['russian']!r}, "
                  f"en={resolved['english']!r}, cat={resolved['category']!r}")
            if resolved["english"]:
                df[resolved["english"]] = (
                    df[resolved["english"]].astype(str)
                    .str.replace(r"\bDoctors of Sciences\b",  "National Doctor", regex=True)
                    .str.replace(r"\bDoctor of Sciences\b",   "National Doctor", regex=True)
                    .str.replace(r"\bCandidates of Sciences\b", "Ph.D.", regex=True)
                    .str.replace(r"\bCandidate of Sciences\b",  "Ph.D.", regex=True)
                )
                df[resolved["english"]] = df[resolved["english"]].where(
                    df[resolved["english"]] != "nan", None)
            rows = []
            synth = 0
            for i, r in df.iterrows():
                gid = parse_int(r.get(resolved["id"])) if resolved["id"] else None
                if gid is None:
                    # Synthesise a stable id from a high offset so it cannot
                    # collide with workbook-supplied #s.  This guarantees that
                    # category-only or partial rows still reach the UI.
                    gid = 1_000_000 + i
                    synth += 1
                ru = s(r.get(resolved["russian"])) if resolved["russian"] else None
                en = s(r.get(resolved["english"])) if resolved["english"] else None
                cat = s(r.get(resolved["category"])) if resolved["category"] else None
                if ru is None and en is None and cat is None:
                    continue  # skip wholly empty rows
                rows.append((gid, ru, en, cat))
            self.conn.executemany(
                "INSERT OR REPLACE INTO glossary VALUES(?,?,?,?)", rows)
            self.stats["glossary"] = len(rows)
            if not rows:
                print("     [warn] Translation Glossary sheet was found but "
                      "produced 0 rows — check column headers "
                      f"(seen: {list(df.columns)}).")
            elif synth:
                print(f"     [info] glossary: synthesised id for {synth} row(s) "
                      "with a blank '#' column.")

        # SheetsIndex — accept either historical name.  The v3 workbook
        # uses "Source Files Index" (with an extra Institution column);
        # earlier exports used "Sheets Index".  These rows describe the
        # primary archival files (ARAN F.1959) that the aggregate workbook
        # was originally compiled from.
        idx_sheet = None
        for cand in ("Sheets Index", "Source Files Index", "Source File Index"):
            if cand in xl:
                idx_sheet = cand; break
        if idx_sheet is not None:
            df = xl[idx_sheet]; rows = []
            for _, r in df.iterrows():
                rows.append((parse_int(r.get("Year")),
                             s(r.get("Institution")),
                             s(r.get("Source File")),
                             s(r.get("Sheet Name")),
                             s(r.get("First Cell (Russian)")),
                             s(r.get("First Cell (English)")),
                             parse_int(r.get("Rows")),
                             parse_int(r.get("Columns")),
                             "primary archival source",
                             None))
            self.conn.executemany(
                "INSERT INTO sheets_index(year, institution, source_file, sheet_name, "
                "first_cell_ru, first_cell_en, rows, cols, role, target_table) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)", rows)
            self.stats["sheets_index"] = len(rows)
            print(f"     [info] sheets_index: loaded {len(rows)} archival "
                  f"provenance rows from sheet '{idx_sheet}'.")

        for k in ["demo_personnel_totals", "demo_academic_degrees", "demo_position",
                  "demo_nationality", "demo_age", "demo_field", "demo_party",
                  "demo_trainees", "glossary", "sheets_index"]:
            if k in self.stats:
                print(f"     {k:24s} {self.stats[k]:>8,}")
        self.conn.commit()

    # ── Phase 2.5: per-input-workbook provenance ────────────────────────

    # Hand-curated mapping from each well-known sheet name to the DB
    # table(s) it feeds.  Sheets that are README / metadata only map to
    # None.  Anything not in this map is recorded with target_table=None
    # but still gets a provenance row so the workbook coverage is total.
    _PROV_TARGETS = {
        # cemi_translation_db.xlsx
        "Personnel Totals":         "demo_personnel_totals",
        "Positions":                "demo_position",
        "Academic Degrees":         "demo_academic_degrees",
        "Nationality":              "demo_nationality",
        "Age Distribution":         "demo_age",
        "Party Membership":         "demo_party",
        "Research Trainees":        "demo_trainees",
        "Translation Glossary":     "glossary",
        "Source Files Index":       "sheets_index (archival rows)",
        "Sheets Index":             "sheets_index (archival rows)",
        "Institution Timeline":     None,
        "README":                   None,
        # data.xlsx
        "CEMI Personnel Data":      ("person, person_year_observation, "
                                     "cemi_appointment, dual_position, "
                                     "pre_cemi_role, transfer_event, "
                                     "dismissal_event, party_event, "
                                     "phd_defense, start_of_work, "
                                     "academic_title_award"),
        "_README":                  None,
        "_Changes_Log":             None,
    }

    def load_pipeline_provenance(self, folder, role: str,
                                  target_table: str = None) -> int:
        """Insert one sheets_index row per sheet in every .xlsx in
        ``folder``.  Used for the OCR Results / English Translation
        Results sibling folders that feed `data.xlsx`.

        The archive file number (e.g. 111 in
        ``cemi_1959_1_111_1966_full.xlsx``) and the year embedded in
        the filename are extracted and recorded — the file number maps
        each OCR/translation file back to its archival counterpart
        (``1959_1_111_отчет.xlsx``) listed in the *Source Files Index*.
        """
        import openpyxl, re, os
        if folder is None or not os.path.isdir(folder):
            print(f"     [info] skipping pipeline provenance for "
                  f"role='{role}': folder not found ({folder!r}).")
            return 0
        n_total = 0
        n_files = 0
        files = sorted(f for f in os.listdir(folder)
                       if f.lower().endswith('.xlsx') and not f.startswith('~$')
                       and not f.startswith('.'))
        for fname in files:
            full = os.path.join(folder, fname)
            # Extract embedded year + archival-file number.
            # OCR/EN naming convention: cemi_1959_1_{archnum}_{year}{tag}.xlsx
            year = None
            archnum = None
            m = re.match(r"cemi_1959_1_(\d+)_(\d{4})", fname)
            if m:
                archnum = m.group(1)
                year = int(m.group(2))
            else:
                # Archival Russian originals: "1959 1 111 отчет.xlsx"
                m2 = re.match(r"1959[\s_]+1[\s_]+(\d+)[\s_]+отчет\.xlsx$", fname)
                if m2:
                    archnum = m2.group(1)
            try:
                wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
            except Exception as e:
                print(f"     [warn] could not open {fname}: {e}")
                continue
            n_files += 1
            rows_out = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                n_rows = ws.max_row or 0
                n_cols = ws.max_column or 0
                a1 = ws["A1"].value if (n_rows and n_cols) else None
                a1 = str(a1).strip() if a1 is not None else None
                # Heuristic: English translations → first_cell_en;
                # everything else (OCR Russian, archival Russian) → first_cell_ru.
                if role == "English translation":
                    ru, en = None, a1
                else:
                    ru, en = a1, None
                # Hand-curated target_table per file (default = personnel pipeline)
                tt = target_table
                if tt is None:
                    tt = ("data.xlsx → person, person_year_observation, "
                          "cemi_appointment, dual_position, pre_cemi_role, "
                          "transfer_event, dismissal_event, party_event, "
                          "phd_defense, start_of_work, academic_title_award")
                # If filename suggests it's a meta-comparison or _README, mark None
                low = fname.lower()
                if (low.startswith('data_corrected_vs') or
                    low.startswith('data_corrected') or
                    low == 'data.xlsx' or
                    'mongodb' in low):
                    tt = ("intermediate consolidation of data.xlsx "
                          "(career data pipeline)")
                inst = ("Central Economic-Mathematical Institute (CEMI) and "
                        "predecessor lab — ARAN F.1959/Op.1") if archnum else None
                # Stamp the archnum into the sheet_name so the trail back to the
                # archival source (`1959_1_<archnum>_отчет.xlsx`) is explicit.
                disp_sheet = sn if not archnum else f"{sn}  [→ 1959_1_{archnum}_отчет.xlsx]"
                rows_out.append((year, inst, fname, disp_sheet, ru, en,
                                 n_rows, n_cols, role, tt))
            self.conn.executemany(
                "INSERT INTO sheets_index(year, institution, source_file, sheet_name, "
                "first_cell_ru, first_cell_en, rows, cols, role, target_table) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)", rows_out)
            n_total += len(rows_out)
        self.conn.commit()
        self.stats["sheets_index"] = self.stats.get("sheets_index", 0) + n_total
        print(f"     [info] sheets_index: added {n_total} rows across "
              f"{n_files} workbooks in '{os.path.basename(folder)}' "
              f"(role='{role}').")
        return n_total

    def load_institution_glossary(self, path) -> None:
        """Ingest the 'institution glossary (Nolting 4-sector …)'
        workbook.  Pure additive — silently no-ops if the file is
        missing so the build still works on lean checkouts.
        """
        import openpyxl, pandas as pd, os
        if path is None or not os.path.exists(str(path)):
            print(f"     [info] institution glossary: file not found "
                  f"({path!r}); skipping.")
            return
        print(f"\n   Loading institution glossary  ←  {os.path.basename(str(path))}")
        try:
            xl = pd.read_excel(str(path), sheet_name=None)
        except Exception as e:
            print(f"     [warn] could not read institution glossary: {e}")
            return

        # (1) Type Codebook → glossary table  ───────────────────────────
        cb = xl.get("Type Codebook")
        n_cb = 0
        if cb is not None:
            # The 5 sector rows carry: Code / Russian / EN definition / KO definition
            for i, r in cb.iterrows():
                code = norm_text(r.get("Code"))
                if not code or code == "—":
                    continue
                ru = norm_text(r.get("Russian"))
                en = norm_text(r.get("Conceptual definition (EN)"))
                # Synthetic stable id well above any workbook-supplied #.
                gid = 2_000_000 + i
                self.conn.execute(
                    "INSERT OR REPLACE INTO glossary VALUES(?,?,?,?)",
                    (gid, ru, f"{code} — {en}" if en else code, "institution_type"))
                n_cb += 1
            self.stats["glossary (institution_type)"] = n_cb
            print(f"     glossary +institution_type    {n_cb:>3}  (Type Codebook entries)")

        # (2) Sheet1/2/3 → back-fill institution.label_ru + evidence +
        #     also accumulate distinct (RU, EN, Type) per institution for
        #     a per-row glossary entry (category='institution').
        n_ru = n_ev = n_kind = 0
        inst_glossary = {}    # english → (russian, type)  — dedupe across sheets
        for sn in ("Sheet1", "Sheet2", "Sheet3"):
            df = xl.get(sn)
            if df is None:
                continue
            # Find the EN / RU / Type / Evidence columns dynamically — the
            # column names differ across sheets ("Previous Institution(English)"
            # vs "Part-Time Institution(English)" etc.).
            # Column detection tolerant of the "Instititution" typo in Sheet1:
            # match by the (english) / russian markers only, ignore the prefix.
            en_col = next((c for c in df.columns
                            if "(english)" in c.lower()), None)
            ru_full_col = next((c for c in df.columns
                                 if "russian full" in c.lower()), None)
            ru_abb_col = next((c for c in df.columns
                                if "russian abb" in c.lower()), None)
            type_col   = "Type" if "Type" in df.columns else None
            ev_col     = next((c for c in df.columns
                                if "Evidence" in c or "근거" in c), None)
            if not en_col:
                continue
            for _, row in df.iterrows():
                en = norm_text(row.get(en_col))
                if not en:
                    continue
                # Reuse the existing institution-id; create one if missing.
                inst_id = self.ensure_inst(en, kind=norm_text(row.get(type_col)))
                ru_full = norm_text(row.get(ru_full_col)) if ru_full_col else None
                ru_abb  = norm_text(row.get(ru_abb_col))  if ru_abb_col  else None
                ru = ru_full or ru_abb     # prefer the full form
                ev = norm_text(row.get(ev_col)) if ev_col else None
                kd = norm_text(row.get(type_col)) if type_col else None
                # COALESCE so a row that lacks a piece does not blank an
                # existing value.
                self.conn.execute(
                    """UPDATE institution
                          SET label_ru               = COALESCE(?, label_ru),
                              classification_evidence = COALESCE(?, classification_evidence),
                              kind                    = COALESCE(?, kind)
                        WHERE institution_id = ?""",
                    (ru, ev, kd, inst_id))
                if ru: n_ru   += 1
                if ev: n_ev   += 1
                if kd: n_kind += 1
                # Capture for the per-row glossary entry.  Dedupe by
                # English name; let the more-informative Russian form win
                # when the same institution appears in multiple sheets.
                if en not in inst_glossary:
                    inst_glossary[en] = (ru, kd)
                else:
                    prev_ru, prev_kd = inst_glossary[en]
                    inst_glossary[en] = (ru or prev_ru, kd or prev_kd)
        self.stats["institution.label_ru"]               = n_ru
        self.stats["institution.classification_evidence"] = n_ev
        print(f"     institution.label_ru          {n_ru:>5}  (rows enriched)")
        print(f"     institution.classific.evidence {n_ev:>5}")

        # (3) Distinct institutions → glossary table (category='institution')
        #     Russian is the canonical RU label, English suffixes the kind
        #     in brackets so the sector is visible from the Glossary tab.
        n_inst = 0
        for i, (en, (ru, kd)) in enumerate(sorted(inst_glossary.items())):
            if not en:
                continue
            gid = 3_000_000 + i      # synthetic id band, no collisions
            display_en = f"{en} [{kd}]" if kd else en
            self.conn.execute(
                "INSERT OR REPLACE INTO glossary VALUES(?,?,?,?)",
                (gid, ru, display_en, "institution"))
            n_inst += 1
        self.stats["glossary (institution)"] = n_inst
        print(f"     glossary +institution         {n_inst:>5}  (distinct from Sheet1/2/3)")
        self.conn.commit()

    def load_workbook_provenance(self, workbook_path, role: str) -> int:
        """Insert one sheets_index row per sheet in ``workbook_path``.

        Uses openpyxl so we get the real (rows, cols) and the A1 cell
        without re-parsing through pandas.  YYYY.MM sheets in
        research_field_subfield.xlsx are recognised and tagged with their
        year and target_table='demo_subfield'.
        """
        import openpyxl, re, os
        fname = os.path.basename(str(workbook_path))
        try:
            wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"     [warn] could not open {fname} for provenance: {e}")
            return 0
        rows_out = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0
            a1 = ws["A1"].value if (n_rows and n_cols) else None
            a1 = str(a1).strip() if a1 is not None else None
            tgt = self._PROV_TARGETS.get(sn, None)
            year = None
            # YYYY.MM sheets in the research workbook
            m = re.fullmatch(r"(\d{4})(?:\.\d{1,2})?", sn.strip())
            if m:
                year = int(m.group(1))
                if tgt is None:
                    tgt = "demo_subfield (+ research_field, research_subfield)"
            rows_out.append((year, None, fname, sn, None, a1, n_rows, n_cols, role, tgt))
        self.conn.executemany(
            "INSERT INTO sheets_index(year, institution, source_file, sheet_name, "
            "first_cell_ru, first_cell_en, rows, cols, role, target_table) "
            "VALUES(?,?,?,?,?,?,?,?,?,?)", rows_out)
        self.conn.commit()
        self.stats["sheets_index"] = self.stats.get("sheets_index", 0) + len(rows_out)
        print(f"     [info] sheets_index: added {len(rows_out)} rows for "
              f"workbook '{fname}' (role='{role}').")
        return len(rows_out)

    # ── Phase 3: subfields ───────────────────────────────────────────────

    def load_subfields(self, xl: dict[str, pd.DataFrame]) -> None:
        """Load the per-year subfield distribution.

        Sheet names are now plain ``YYYY`` (e.g. ``1965``, ``1971``).
        The ``period`` column simply mirrors the sheet name, so it
        equals ``"YYYY"`` for the current workbook.  Older workbooks
        that used ``YYYY.MM`` snapshot labels (e.g. ``1971.01`` /
        ``1971.06``) are still accepted: the regex below picks them up
        and the raw label is preserved in ``period``.  This keeps the
        ``(year, period, subfield_id)`` primary key collision-free in
        either layout.
        """
        import re as _re
        print("\n   Loading subfield hierarchy")
        import json as _json
        n_field = n_sub = n_fact = 0
        for sheet_name, df in xl.items():
            name = str(sheet_name).strip()
            m = _re.fullmatch(r"(\d{4})(?:\.(\d{1,2}))?", name)
            if not m:
                # Skip non-period sheets (legacy "Notes", "Index" etc.)
                continue
            year = int(m.group(1))
            # When the sheet name carries a month suffix (legacy v3
            # workbooks), preserve it.  Otherwise period == "YYYY".
            period = name if m.group(2) else str(year)
            if "Research Field" not in df.columns:
                continue
            self.ensure_year(year)
            signature = "|".join(str(c) for c in df.columns)
            df = df.where(pd.notna(df), None)
            current_field: str | None = None

            for _, r in df.iterrows():
                fld = norm_text(r["Research Field"])
                sub = norm_text(r.get("Subfield"))
                if fld:
                    current_field = fld
                    fid = self.rfield.get(fld, code=None, label_ru=None, is_aggregate=1)
                    n_field += 1
                if not sub or sub.lower() in {"total", "итого"}:
                    continue
                fid = self.rfield.get(current_field or "Unknown", code=None,
                                      label_ru=None, is_aggregate=1)
                sub_id = self.rsub.get(sub, rfield_id=fid)
                n_sub += 1
                extras: dict[str, Any] = {}
                for col, val in r.items():
                    if col in {"Research Field", "Subfield",
                               "Total Personnel", "National Doctor", "Ph.D."}:
                        continue
                    if val is not None and not (isinstance(val, float) and pd.isna(val)):
                        if isinstance(val, (int, float)):
                            extras[str(col)] = float(val)
                        else:
                            extras[str(col)] = str(val)
                self.conn.execute("""
                    INSERT OR REPLACE INTO demo_subfield(year, period, subfield_id,
                        total_personnel, national_doctor, phd, extras_json,
                        schema_year_signature)
                    VALUES(?,?,?,?,?,?,?,?)""",
                    (year, period, sub_id,
                     parse_float(r.get("Total Personnel")),
                     parse_float(r.get("National Doctor")),
                     parse_float(r.get("Ph.D.")),
                     _json.dumps(extras, ensure_ascii=False) if extras else None,
                     signature))
                n_fact += 1
        self.stats["demo_subfield"] = n_fact
        print(f"     research_field           {n_field:>8,}")
        print(f"     research_subfield        {n_sub:>8,}")
        print(f"     demo_subfield            {n_fact:>8,}")
        self.conn.commit()

    # ── Phase 3.5: derive field-level aggregates from subfields ──────────
    #
    # Since cemi_translation_db.xlsx · "Research Fields" sheet has been
    # retired (the user replaced it entirely with research_field_subfield),
    # the yearly Field totals that used to live in `demo_field` are now
    # *derived* — by summing the per-year subfields per (year, Research
    # Field).  Every year has a single snapshot now (period == str(year));
    # the MAX(period) selection below remains a defensive no-op that still
    # picks the canonical snapshot if any future workbook reintroduces
    # YYYY.MM labels.

    def derive_field_aggregates(self) -> None:
        print("\n   Deriving demo_field from subfield aggregation")
        cur = self.conn.cursor()
        cur.execute("DELETE FROM demo_field")    # always start clean
        # Latest snapshot per year.
        latest = {row[0]: row[1] for row in
                  cur.execute("SELECT year, MAX(period) FROM demo_subfield GROUP BY year").fetchall()}
        rows = []
        for year, period in latest.items():
            for r in cur.execute("""
                SELECT rs.rfield_id,
                       SUM(ds.total_personnel),
                       SUM(ds.national_doctor),
                       SUM(ds.phd)
                  FROM demo_subfield ds
                  JOIN research_subfield rs ON rs.subfield_id = ds.subfield_id
                 WHERE ds.year = ? AND ds.period = ?
                 GROUP BY rs.rfield_id""", (year, period)).fetchall():
                tot = int(r[1]) if r[1] is not None else None
                rows.append((year, r[0], tot, r[2], r[3]))
        cur.executemany("INSERT INTO demo_field VALUES(?,?,?,?,?)", rows)
        self.conn.commit()
        self.stats["demo_field (derived)"] = len(rows)
        print(f"     demo_field rows derived:  {len(rows):>5}  "
              f"(from {len(latest)} year-snapshots)")

    # ── Phase 4: validate ─────────────────────────────────────────────────

    def validate(self) -> None:
        print("\n   Validation")
        cur = self.conn.cursor()
        ic = cur.execute("PRAGMA integrity_check;").fetchone()[0]
        fk = cur.execute("PRAGMA foreign_key_check;").fetchall()
        print(f"     integrity_check  : {ic}")
        print(f"     foreign_key_check: {len(fk)} violation(s)")
        # Show row counts for the career-side tables (most important)
        career_tables = [
            "person", "person_name_variant", "person_year_observation",
            "cemi_appointment", "dual_position", "pre_cemi_role",
            "transfer_event", "dismissal_event", "party_event",
            "phd_defense", "start_of_work", "academic_title_award",
        ]
        print("\n   Career-side tables")
        for t in career_tables:
            n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            print(f"     {t:24s} {n:>8,}")

    def close(self) -> None:
        self.conn.commit()
        self.conn.close()


# ══════════════════════════════════════════════════════════════════════════════
# §F   File discovery (P6: runs from anywhere)
# ══════════════════════════════════════════════════════════════════════════════

def script_directory() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd().resolve()


def locate(user_value: str, default_name: str, label: str) -> Path:
    cwd  = Path.cwd().resolve()
    here = script_directory()
    candidates: list[Path] = []
    if user_value:
        p = Path(user_value).expanduser()
        candidates.append(p if p.is_absolute() else (cwd / p))
        if not p.is_absolute() and p.parent == Path("."):
            candidates.append(here / p.name)
    candidates.append(cwd  / default_name)
    candidates.append(here / default_name)

    seen, ordered = set(), []
    for c in candidates:
        r = c.resolve()
        if r in seen: continue
        seen.add(r); ordered.append(r)
    for c in ordered:
        if c.exists():
            return c
    print(f"[ERROR] {label} ({default_name}) not found.  Searched:")
    for c in ordered: print(f"          - {c}")
    sys.exit(1)


def resolve_db(user_value: str, default_name: str) -> Path:
    p = Path(user_value or default_name).expanduser()
    return (p if p.is_absolute() else (Path.cwd() / p)).resolve()


# ══════════════════════════════════════════════════════════════════════════════
# §G   CLI
# ══════════════════════════════════════════════════════════════════════════════

DEFAULTS = {
    "personnel":   "data.xlsx",
    "translation": "cemi_translation_db.xlsx",
    "research":    "research_field_subfield.xlsx",
    "db":          "cemi_career.db",
}


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--personnel",   default=DEFAULTS["personnel"])
    p.add_argument("--translation", default=DEFAULTS["translation"])
    p.add_argument("--research",    default=DEFAULTS["research"])
    p.add_argument("--db",          default=DEFAULTS["db"])
    # `--reset` is now a no-op kept only for backward compatibility — the
    # builder always deletes the existing DB and rebuilds from scratch so
    # that every run produces a deterministic, clean output.
    p.add_argument("--reset", action="store_true",
                   help="(no-op; the builder always overwrites the existing DB).")
    args = p.parse_args()

    pers = locate(args.personnel,   DEFAULTS["personnel"],   "personnel")
    tran = locate(args.translation, DEFAULTS["translation"], "translation")
    rese = locate(args.research,    DEFAULTS["research"],    "research")
    dbp  = resolve_db(args.db,      DEFAULTS["db"])

    # Auto-discover the institution glossary workbook in the same folder
    # as data.xlsx.  Match the canonical filename ("institution glossary*.xlsx")
    # case-insensitively.
    import glob as _glob, os as _os
    _hits = _glob.glob(_os.path.join(_os.path.dirname(str(pers)),
                                       "institution glossary*.xlsx"))
    inst_gloss = _hits[0] if _hits else None

    print("── Resolved inputs ──")
    print(f"   personnel   : {pers}")
    print(f"   translation : {tran}")
    print(f"   research    : {rese}")
    print(f"   output DB   : {dbp}")

    dbp.parent.mkdir(parents=True, exist_ok=True)
    # Always overwrite — a stale DB file from a previous build would
    # otherwise blend with the new run and confuse downstream tools.
    if dbp.exists():
        dbp.unlink()
        print("[OK] Removed previous DB file (auto-overwrite)")

    stage = Stage(dbp)

    df_pers = pd.read_excel(pers, sheet_name="CEMI Personnel Data")
    stage.load_personnel(df_pers)
    stage.derive_career_artifacts()

    xl_tran = pd.read_excel(tran, sheet_name=None)
    stage.load_aggregates(xl_tran)
    # Institution glossary (Nolting 4-sector classification) — adds RU
    # labels + classification evidence to the institution table and adds
    # 5 institution_type entries to the glossary table.
    stage.load_institution_glossary(inst_gloss)

    xl_rese = pd.read_excel(rese, sheet_name=None)
    stage.load_subfields(xl_rese)
    stage.derive_field_aggregates()

    # Per-input-workbook provenance — one sheets_index row per worksheet
    # in each of the three input workbooks.  The 147 archival rows from
    # "Source Files Index" are kept alongside (role='primary archival source').
    print("\n   Recording per-input-workbook provenance")
    stage.load_workbook_provenance(pers, role="personnel workbook")
    stage.load_workbook_provenance(tran, role="intermediate workbook")
    stage.load_workbook_provenance(rese, role="intermediate workbook")

    # Pipeline provenance — every Excel workbook that sits upstream of
    # the final aggregate / personnel workbooks.  The folders live
    # alongside data.xlsx in the project tree.  Missing folders are
    # silently no-op'd so the build still works on lean checkouts.
    print("\n   Recording pipeline provenance (career + demographic)")
    base_dir = pers.parent

    # Career-side pipeline
    ocr_dir  = base_dir / "OCR Results"
    en_dir   = base_dir / "English Tranlation Results"   # folder has a typo
    if not en_dir.is_dir():
        alt = base_dir / "English Translation Results"
        if alt.is_dir():
            en_dir = alt
    stage.load_pipeline_provenance(ocr_dir, role="OCR intermediate")
    stage.load_pipeline_provenance(en_dir,  role="English translation")

    # Demographic-side pipeline — the 23 Russian archival originals
    # (1959 1 X отчет.xlsx) saved locally as Excel.  These are the
    # actual Excel files that the "Source Files Index" sheet of
    # cemi_translation_db.xlsx describes as 147 metadata rows.
    demog_dir = base_dir / "demographic db(new)"
    stage.load_pipeline_provenance(
        demog_dir,
        role="archival workbook (Russian original)",
        target_table=(
            "cemi_translation_db.xlsx aggregation → demo_personnel_totals, "
            "demo_position, demo_academic_degrees, demo_nationality, demo_age, "
            "demo_party, demo_trainees, demo_field"))

    stage.validate()
    stage.close()
    print(f"\n[OK] Built {dbp}")


if __name__ == "__main__":
    main()
